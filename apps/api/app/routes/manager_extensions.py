"""Manager-facing compatibility endpoints for the current UI mock flow.

These endpoints deliberately keep the existing numeric identifiers and domain
statuses.  The mock UI can use the extra display fields while the core
schedule state machine remains unchanged.
"""

from __future__ import annotations

import re
import secrets
from datetime import UTC, date, datetime
from io import BytesIO
from typing import Annotated, Any, Literal
from zoneinfo import ZoneInfo

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, ConfigDict, Field
from sqlalchemy import text
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.api_contract import parse_external_id, success_payload
from app.auth import CurrentUser, get_current_user
from app.config import get_settings
from app.database import get_db
from app.domain.errors import DomainError
from app.domain.master_data import normalize_code
from app.domain.registration_phase import resolve_registration_phase
from app.domain.round_setup import validate_round_configuration
from app.response_models import (
    ActionResponse,
    GroupDetailResponse,
    GroupResponse,
    ImportResponse,
    InvitationResponse,
    LecturerImportResponse,
    ProjectDetailResponse,
    ProjectMutationResponse,
    QuotaResponse,
    RescheduleRequestResponse,
    ResultResponse,
    RoundDetailEnvelopeResponse,
    RoundGroupResponse,
    SemesterResponse,
    SessionResponse,
    TimeslotResponse,
)
from app.routes.master_data import _insert_timeframe_slots, password_hasher
from app.services.semester_queries import (
    SEMESTER_LIFECYCLE_LOCK_KEY,
    academic_year_for_start,
    ensure_round_semester_writable,
    ensure_semester_writable,
    semester_or_404,
)
from app.services.timeframe_service import load_timeframe_template

_PHASE3_PROVENANCE_TABLE = "schedule_assignments"
router = APIRouter(prefix="/api/v1", tags=["manager-ui-compatibility"])
Db = Annotated[Session, Depends(get_db)]
User = Annotated[CurrentUser, Depends(get_current_user)]


def _require(user: CurrentUser, *roles: str) -> None:
    if user.role not in roles:
        raise HTTPException(status_code=403, detail="Insufficient permission")


def _actor_id(db: Session, user: CurrentUser) -> int | None:
    if user.account_id is not None:
        return user.account_id
    return db.execute(
        text("SELECT MIN(account_id) FROM account_roles WHERE role = CAST(:role AS system_role)"),
        {"role": user.role},
    ).scalar_one_or_none()


def _json(value: Any) -> str:
    import json

    return json.dumps(value, default=str)


class ProjectUpdate(BaseModel):
    model_config = ConfigDict(populate_by_name=True)
    code: str | None = Field(default=None, min_length=1, max_length=64)
    title: str | None = Field(default=None, min_length=1, max_length=255)
    name_vi: str | None = Field(default=None, alias="nameVi", min_length=1, max_length=255)
    name_en: str | None = Field(default=None, alias="nameEn", max_length=255)
    main_supervisor_id: str | int | None = Field(default=None, alias="mainSupervisorId")
    co_supervisor_id: str | int | None = Field(default=None, alias="coSupervisorId")
    supervisors: list[str] | None = None


class GroupUpdate(BaseModel):
    code: str | None = Field(default=None, min_length=1, max_length=64)
    project_id: int | None = Field(
        default=None, gt=0, description="Gán/đổi project cho nhóm. Truyền null để bỏ gán."
    )


class RoundUpdate(BaseModel):
    model_config = ConfigDict(populate_by_name=True)

    start_date: date | None = Field(default=None, alias="startDate")
    end_date: date | None = Field(default=None, alias="endDate")
    session_duration_minutes: int | None = Field(default=None, alias="durationMinutes", gt=0, le=480)
    reviewer_count: int | None = Field(default=None, alias="reviewerCount", gt=0)
    result_owner_mode: bool | None = Field(default=None, alias="resultOwnerMode")
    group_selection_mode: bool | None = Field(default=None, alias="groupSelectionMode")
    registration_deadline: datetime | None = Field(default=None, alias="registrationDeadline")
    group_preference_deadline: datetime | None = Field(default=None, alias="groupPreferenceDeadline")
    h12_sessions_per_part: int | None = Field(default=None, alias="h12SessionsPerPart", gt=0)
    h12_sessions_per_day: int | None = Field(default=None, alias="h12SessionsPerDay", gt=0)
    h12_semester_quota: int | None = Field(default=None, alias="h12SemesterQuota", gt=0)
    max_groups_per_timeslot: int | None = Field(default=None, alias="maxGroupsPerTimeslot", gt=0)
    max_minutes_per_part: int | None = Field(default=None, alias="maxMinutesPerPart", gt=0)
    max_minutes_per_day: int | None = Field(default=None, alias="maxMinutesPerDay", gt=0)
    room_types: list[Literal["NORMAL", "SEMINAR", "LAB"]] | None = Field(default=None, alias="roomTypes")
    timeframe_id: int | None = Field(default=None, alias="timeframeId", gt=0)


class TimeslotUpdate(BaseModel):
    start_at: datetime | None = None
    end_at: datetime | None = None
    active: bool | None = None
    part: str | None = Field(default=None, pattern="^(AM|PM)$")


class QuotaUpdate(BaseModel):
    quota: int = Field(gt=0)


class SemesterUpdate(BaseModel):
    code: str | None = Field(default=None, min_length=1, max_length=32)
    name: str | None = Field(default=None, min_length=1, max_length=160)
    note: str | None = Field(default=None, max_length=1000)
    start_date: date | None = None
    end_date: date | None = None


def _row_with_json(row: Any) -> dict[str, Any]:
    return dict(row)


@router.patch("/semesters/{semester_id}", response_model=SemesterResponse)
def update_semester(semester_id: int, payload: SemesterUpdate, db: Db, user: User) -> dict[str, Any]:
    _require(user, "ADMIN", "MANAGER")
    current = db.execute(
        text(
            "SELECT id, code, name, note, start_date, end_date, academic_year, status, "
            "created_by, updated_by, created_at, updated_at FROM semesters WHERE id = :id"
        ),
        {"id": semester_id},
    ).mappings().one_or_none()
    db.rollback()
    if current is None:
        raise HTTPException(status_code=404, detail={"code": "SEMESTER_NOT_FOUND", "message": "Semester does not exist."})
    incoming = payload.model_dump(exclude_unset=True)
    values = {**dict(current), **incoming}
    if "code" in values and values["code"] is not None:
        values["code"] = normalize_code(values["code"])
    if "name" in values and values["name"] is not None:
        values["name"] = values["name"].strip()
    if "note" in values and values["note"] is not None:
        values["note"] = values["note"].strip() or None
    if values["start_date"] is None or values["end_date"] is None:
        raise HTTPException(status_code=422, detail={"code": "SEMESTER_DATE_INVALID", "message": "Both semester dates are required."})
    if values["end_date"] < values["start_date"]:
        raise HTTPException(status_code=422, detail={"code": "SEMESTER_DATE_INVALID", "message": "end_date must be after start_date."})
    settings = get_settings()
    inclusive_days = (values["end_date"] - values["start_date"]).days + 1
    if not settings.semester_min_duration_days <= inclusive_days <= settings.semester_max_duration_days:
        raise HTTPException(status_code=422, detail={"code": "SEMESTER_DURATION_INVALID", "message": "Semester duration is outside the configured range."})
    changed = {key: values[key] for key in incoming}
    if not changed:
        return semester_or_404(db, semester_id)
    changed["academic_year"] = academic_year_for_start(values["start_date"])
    try:
        with db.begin():
            db.execute(
                text("SELECT pg_advisory_xact_lock(:lock_key)"),
                {"lock_key": SEMESTER_LIFECYCLE_LOCK_KEY},
            )
            locked = db.execute(
                text(
                    "SELECT code, name, note, start_date, end_date, academic_year, status "
                    "FROM semesters WHERE id = :id FOR UPDATE"
                ),
                {"id": semester_id},
            ).mappings().one_or_none()
            if locked is None:
                raise HTTPException(status_code=404, detail={"code": "SEMESTER_NOT_FOUND", "message": "Semester does not exist."})
            if str(locked["status"]) == "ARCHIVED":
                raise HTTPException(status_code=409, detail={"code": "SEMESTER_ARCHIVED", "message": "An archived semester is read-only."})
            actor_id = _actor_id(db, user)
            db.execute(
                text(
                    "UPDATE semesters SET code = :code, name = :name, note = :note, "
                    "start_date = :start_date, end_date = :end_date, academic_year = :academic_year, "
                    "updated_by = :updated_by, updated_at = now() WHERE id = :id"
                ),
                {
                    "code": values["code"],
                    "name": values["name"],
                    "note": values.get("note"),
                    "start_date": values["start_date"],
                    "end_date": values["end_date"],
                    "academic_year": changed["academic_year"],
                    "id": semester_id,
                    "updated_by": actor_id,
                },
            )
            db.execute(
                text(
                    "INSERT INTO audit_events "
                    "(actor_id, action, entity_type, entity_id, before_json, after_json) "
                    "VALUES (:actor_id, 'SEMESTER_UPDATED', 'semester', :entity_id, "
                    "CAST(:before_json AS JSONB), CAST(:after_json AS JSONB))"
                ),
                {
                    "actor_id": actor_id,
                    "entity_id": str(semester_id),
                    "before_json": _json(dict(locked)),
                    "after_json": _json({**dict(locked), **changed}),
                },
            )
    except IntegrityError as exc:
        raise HTTPException(
            status_code=409,
            detail={"code": "DATA_DUPLICATE", "message": "Semester code already exists."},
        ) from exc
    return semester_or_404(db, semester_id)


def _load_round_detail(round_id: int, db: Session) -> dict[str, Any]:
    row = db.execute(
        text(
            """
            SELECT r.*, s.code AS semester_code, s.name AS semester_name,
                   COALESCE((SELECT MIN(day_date) FROM round_days WHERE round_id = r.id), r.start_date) AS effective_start_date,
                   COALESCE((SELECT MAX(day_date) FROM round_days WHERE round_id = r.id), r.end_date) AS effective_end_date,
                   (SELECT COUNT(*) FROM round_groups WHERE round_id = r.id) AS group_count,
                   (SELECT COUNT(*) FROM rooms rm JOIN round_room_types rrt ON rrt.room_type = rm.room_type
                    WHERE rrt.round_id = r.id AND rm.active = TRUE) AS room_count,
                   COALESCE((SELECT array_agg(rrt.room_type::text ORDER BY rrt.room_type::text)
                             FROM round_room_types rrt WHERE rrt.round_id = r.id), ARRAY[]::text[]) AS room_types,
                   (SELECT COUNT(*) FROM timeslots ts JOIN round_days rd ON rd.id = ts.round_day_id WHERE rd.round_id = r.id AND ts.active) AS active_timeslot_count
            FROM rounds r JOIN semesters s ON s.id = r.semester_id WHERE r.id = :id
            """
        ),
        {"id": round_id},
    ).mappings().one_or_none()
    if row is None:
        raise HTTPException(status_code=404, detail={"code": "ROUND_NOT_FOUND", "message": "Round does not exist."})
    days = db.execute(
        text(
            "SELECT rd.id, rd.day_date, ts.id AS timeslot_id, ts.start_at, ts.end_at, ts.part, ts.active "
            "FROM round_days rd LEFT JOIN timeslots ts ON ts.round_day_id = rd.id WHERE rd.round_id = :id "
            "ORDER BY rd.day_date, ts.start_at"
        ),
        {"id": round_id},
    ).mappings().all()
    return {**dict(row), "days": [dict(item) for item in days]}


_ROUND_DETAIL_TIMEZONE = ZoneInfo("Asia/Ho_Chi_Minh")


def _local_time(value: datetime) -> str:
    if value.tzinfo is None:
        value = value.replace(tzinfo=UTC)
    return value.astimezone(_ROUND_DETAIL_TIMEZONE).strftime("%H:%M")


def _round_detail_payload(raw: dict[str, Any]) -> dict[str, Any]:
    days_by_date: dict[date, dict[str, Any]] = {}
    for item in raw["days"]:
        day_date = item["day_date"]
        day = days_by_date.setdefault(day_date, {"date": day_date, "slots": []})
        if item["timeslot_id"] is None or not item["active"]:
            continue
        day["slots"].append(
            {
                "id": str(item["timeslot_id"]),
                "start_time": _local_time(item["start_at"]),
                "end_time": _local_time(item["end_at"]),
            }
        )

    registration_phase = resolve_registration_phase(
        round_status=str(raw["status"]),
        registration_deadline=raw["registration_deadline"],
        group_preference_deadline=raw["group_preference_deadline"],
        lecturer_registration_closed_at=raw.get("lecturer_registration_closed_at"),
        now=datetime.now(UTC),
    )

    return success_payload(
        {
            "id": str(raw["id"]),
            "semester_id": str(raw["semester_id"]),
            "name": raw["name"],
            "type": str(raw["type"]),
            "status": str(raw["status"]),
            "description": raw["description"],
            "duration_minutes": raw["session_duration_minutes"],
            "reviewer_count": raw["reviewer_count"],
            "max_groups_per_timeslot": raw["max_groups_per_timeslot"],
            "registration_deadline": raw["registration_deadline"],
            "group_selection_mode": raw["group_selection_mode"],
            "group_preference_deadline": raw["group_preference_deadline"],
            "registration_phase": registration_phase.value,
            "result_owner_mode": raw["result_owner_mode"],
            "room_types": raw["room_types"],
            "timeframe_id": str(raw["timeframe_id"]) if raw.get("timeframe_id") is not None else None,
            "timeframe_version_id": str(raw["timeframe_version_id"]) if raw.get("timeframe_version_id") is not None else None,
            "days": list(days_by_date.values()),
        }
    )


@router.get(
    "/rounds/{round_id}",
    response_model=RoundDetailEnvelopeResponse,
    response_model_exclude_none=True,
)
def get_round_detail(round_id: int, db: Db, user: User) -> dict[str, Any]:
    _require(user, "ADMIN", "MANAGER")
    return _round_detail_payload(_load_round_detail(round_id, db))


@router.patch("/rounds/{round_id}", response_model=RoundDetailEnvelopeResponse)
def update_round(round_id: int, payload: RoundUpdate, db: Db, user: User) -> dict[str, Any]:
    _require(user, "ADMIN", "MANAGER")
    values = payload.model_dump(exclude_unset=True)
    if not values:
        return _round_detail_payload(_load_round_detail(round_id, db))
    if values.get("start_date") and values.get("end_date") and values["end_date"] < values["start_date"]:
        raise HTTPException(status_code=422, detail={"code": "ROUND_DATE_INVALID", "message": "end_date must be after start_date."})
    columns = {"start_date", "end_date", "session_duration_minutes", "reviewer_count", "result_owner_mode", "group_selection_mode", "registration_deadline", "group_preference_deadline", "h12_sessions_per_part", "h12_sessions_per_day", "h12_semester_quota", "max_groups_per_timeslot", "max_minutes_per_part", "max_minutes_per_day", "timeframe_id"}
    assignments = [f"{key} = :{key}" for key in values if key in columns]
    if not assignments and "room_types" not in values:
        raise HTTPException(status_code=422, detail={"code": "ROUND_UPDATE_EMPTY", "message": "No editable round fields supplied."})
    values["id"] = round_id
    with db.begin():
        ensure_round_semester_writable(db, round_id)
        current_round = db.execute(text("SELECT status, type, reviewer_count, result_owner_mode, group_selection_mode, registration_deadline, group_preference_deadline, start_date, end_date, session_duration_minutes, timeframe_id, timeframe_version_id FROM rounds WHERE id = :id FOR UPDATE"), {"id": round_id}).mappings().one_or_none()
        if current_round is None:
            raise HTTPException(status_code=404, detail={"code": "ROUND_NOT_FOUND", "message": "Round does not exist."})
        if str(current_round["status"]) not in {"DRAFT", "OPEN_REGISTRATION"}:
            raise HTTPException(status_code=409, detail={"code": "ROUND_CONFIG_LOCKED", "message": "Round configuration can only be edited before scheduling."})
        merged_start = values.get("start_date", current_round["start_date"])
        merged_end = values.get("end_date", current_round["end_date"])
        if merged_start is not None and merged_end is not None and merged_end < merged_start:
            raise HTTPException(status_code=422, detail={"code": "ROUND_DATE_INVALID", "message": "end_date must be after start_date."})
        merged_registration_deadline = values.get("registration_deadline", current_round["registration_deadline"])
        merged_group_preference_deadline = values.get("group_preference_deadline", current_round["group_preference_deadline"])
        if merged_registration_deadline is not None and merged_registration_deadline.tzinfo is None:
            raise HTTPException(status_code=422, detail={"code": "ROUND_DEADLINE_INVALID", "message": "registration_deadline must include a timezone offset."})
        if merged_group_preference_deadline is not None and merged_group_preference_deadline.tzinfo is None:
            raise HTTPException(status_code=422, detail={"code": "ROUND_DEADLINE_INVALID", "message": "group_preference_deadline must include a timezone offset."})
        if merged_start is not None and merged_end is not None:
            for field_name, deadline in (("registration_deadline", merged_registration_deadline), ("group_preference_deadline", merged_group_preference_deadline)):
                if deadline is not None and not (merged_start <= deadline.date() <= merged_end):
                    raise HTTPException(status_code=422, detail={"code": "ROUND_DEADLINE_INVALID", "message": f"{field_name} must fall within start_date and end_date."})
        requested_timeframe = values.get("timeframe_id", current_round["timeframe_id"])
        timeframe_changed = "timeframe_id" in values and values["timeframe_id"] != current_round["timeframe_id"]
        if "timeframe_id" in values and values["timeframe_id"] is None and current_round["timeframe_id"] is not None:
            raise HTTPException(status_code=409, detail={"code": "ROUND_TIMEFRAME_UNBIND_NOT_ALLOWED", "message": "A Round with generated Timeframe slots cannot be unbound from its Timeframe."})
        timeframe_version_id = current_round["timeframe_version_id"]
        generated_timeframe = None
        regenerate_timeframe = False
        if requested_timeframe is not None:
            if timeframe_changed or current_round["timeframe_id"] is None:
                timeframe_version_id, generated_timeframe = load_timeframe_template(
                    db,
                    timeframe_id=int(requested_timeframe),
                )
            else:
                timeframe_version_id, generated_timeframe = load_timeframe_template(
                    db,
                    timeframe_id=int(requested_timeframe),
                    version_id=int(current_round["timeframe_version_id"]),
                )
            merged_duration = values.get("session_duration_minutes", current_round["session_duration_minutes"])
            if merged_duration != generated_timeframe.group_duration_minutes:
                raise HTTPException(status_code=422, detail={"code": "TIMEFRAME_SESSION_DURATION_MISMATCH", "message": "Round duration must equal the Timeframe group duration."})
            regenerate_timeframe = timeframe_changed or "start_date" in values or "end_date" in values
        if regenerate_timeframe:
            if str(current_round["status"]) != "DRAFT":
                raise HTTPException(status_code=409, detail={"code": "ROUND_TIMEFRAME_LOCKED", "message": "Timeframe-generated slots can only be regenerated while the Round is DRAFT."})
            dependent_slots = db.execute(
                text(
                    "SELECT (SELECT COUNT(*) FROM lecturer_availabilities WHERE round_id = :id) "
                    "+ (SELECT COUNT(*) FROM group_slot_preferences WHERE round_id = :id)"
                ),
                {"id": round_id},
            ).scalar_one()
            if int(dependent_slots) > 0:
                raise HTTPException(status_code=409, detail={"code": "ROUND_TIMEFRAME_REGENERATION_BLOCKED", "message": "Cannot regenerate Timeframe slots after availability or group preferences exist."})
            db.execute(text("DELETE FROM round_days WHERE round_id = :id"), {"id": round_id})
            if generated_timeframe is None:
                raise HTTPException(status_code=422, detail={"code": "TIMEFRAME_NOT_FOUND", "message": "The selected Timeframe does not have a usable revision."})
            merged_start = values.get("start_date", current_round["start_date"])
            merged_end = values.get("end_date", current_round["end_date"])
            _insert_timeframe_slots(
                db,
                round_id=round_id,
                start_date=merged_start,
                end_date=merged_end,
                generated=generated_timeframe,
            )
        if timeframe_changed:
            values["timeframe_version_id"] = timeframe_version_id
            assignments.append("timeframe_version_id = :timeframe_version_id")
        try:
            validate_round_configuration({"type": str(current_round["type"]), "reviewer_count": values.get("reviewer_count", current_round["reviewer_count"]), "result_owner_mode": values.get("result_owner_mode", current_round["result_owner_mode"]), "groups": [1], "timeslots": [1], "rooms": [1]})
        except DomainError as exc:
            raise HTTPException(status_code=422, detail={"code": exc.code, "message": str(exc)}) from exc
        if assignments:
            updated = db.execute(text(f"UPDATE rounds SET {', '.join(assignments)} WHERE id = :id RETURNING id"), values).scalar_one_or_none()
            if updated is None:
                raise HTTPException(status_code=404, detail={"code": "ROUND_NOT_FOUND", "message": "Round does not exist."})
        if "room_types" in values:
            room_types = values["room_types"] or []
            if not room_types:
                raise HTTPException(status_code=422, detail={"code": "ROUND_ROOM_TYPES_REQUIRED", "message": "At least one allowed room type is required."})
            db.execute(text("DELETE FROM round_room_types WHERE round_id = :id"), {"id": round_id})
            for room_type in sorted(set(room_types)):
                db.execute(text("INSERT INTO round_room_types (round_id, room_type) VALUES (:round_id, CAST(:room_type AS room_type))"), {"round_id": round_id, "room_type": room_type})
    return _round_detail_payload(_load_round_detail(round_id, db))


@router.patch("/projects/{project_id}", response_model=ProjectMutationResponse)
def update_project(project_id: str, payload: ProjectUpdate, db: Db, user: User) -> dict[str, Any]:
    _require(user, "ADMIN", "MANAGER")
    try:
        project_id = parse_external_id(project_id, prefix="prj")
    except (TypeError, ValueError) as exc:
        raise HTTPException(
            status_code=422,
            detail={"code": "VALIDATION_ERROR", "message": "project_id must be a valid project identifier."},
        ) from exc
    values = payload.model_dump(exclude_unset=True)
    target_supervisor_update = "main_supervisor_id" in values or "co_supervisor_id" in values
    target_supervisors: list[tuple[int, str]] | None = None
    if target_supervisor_update:
        main_id = values.pop("main_supervisor_id", None)
        co_id = values.pop("co_supervisor_id", None)
        if main_id is None:
            raise HTTPException(status_code=422, detail={"code": "SUPERVISOR_INVALID", "message": "mainSupervisorId is required."})
        try:
            main_db_id = parse_external_id(main_id, prefix="lec")
            co_db_id = parse_external_id(co_id, prefix="lec") if co_id is not None else None
        except (TypeError, ValueError) as exc:
            raise HTTPException(status_code=422, detail={"code": "SUPERVISOR_INVALID", "message": "Supervisor identifier is invalid."}) from exc
        if co_db_id is not None and co_db_id == main_db_id:
            raise HTTPException(status_code=422, detail={"code": "SUPERVISOR_INVALID", "message": "mainSupervisorId and coSupervisorId must differ."})
        target_supervisors = [(main_db_id, "MAIN")]
        if co_db_id is not None:
            target_supervisors.append((co_db_id, "CO"))
    if "name_vi" in values:
        values["title"] = values.pop("name_vi")
    values.pop("name_en", None)
    try:
        with db.begin():
            row = db.execute(text("SELECT id, code, title, semester_id FROM projects WHERE id = :id FOR UPDATE"), {"id": project_id}).mappings().one_or_none()
            if row is None:
                raise HTTPException(status_code=404, detail={"code": "PROJECT_NOT_FOUND", "message": "Project does not exist."})
            ensure_semester_writable(db, int(row["semester_id"]))
            scalar = {key: values[key] for key in ("code", "title") if key in values}
            if scalar:
                assignments = ", ".join(f"{key} = :{key}" for key in scalar)
                scalar["id"] = project_id
                db.execute(text(f"UPDATE projects SET {assignments} WHERE id = :id"), scalar)
            if target_supervisor_update or "supervisors" in values:
                db.execute(text("DELETE FROM project_supervisors WHERE project_id = :id"), {"id": project_id})
                if target_supervisor_update:
                    assignments = target_supervisors or []
                else:
                    assignments = []
                    for assignment in values["supervisors"] or []:
                        code, _, supervisor_type = assignment.partition(":")
                        lecturer_id = db.execute(text("SELECT id FROM lecturers WHERE lecturer_code = :code"), {"code": code.strip().upper()}).scalar_one_or_none()
                        if lecturer_id is None or supervisor_type.strip().upper() not in {"MAIN", "CO"}:
                            raise HTTPException(status_code=422, detail={"code": "SUPERVISOR_INVALID", "message": "Supervisor code/type is invalid."})
                        assignments.append((lecturer_id, supervisor_type.strip().upper()))
                existing_lecturers = db.execute(
                    text("SELECT id FROM lecturers WHERE id = ANY(:ids)"),
                    {"ids": [lecturer_id for lecturer_id, _ in assignments]},
                ).scalars().all()
                if set(existing_lecturers) != {lecturer_id for lecturer_id, _ in assignments}:
                    raise HTTPException(status_code=422, detail={"code": "SUPERVISOR_INVALID", "message": "One or more supervisors do not exist."})
                for lecturer_id, supervisor_type in assignments:
                    db.execute(text("INSERT INTO project_supervisors (project_id, lecturer_id, supervisor_type) VALUES (:project_id, :lecturer_id, CAST(:type AS supervisor_type))"), {"project_id": project_id, "lecturer_id": lecturer_id, "type": supervisor_type})
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(status_code=409, detail={"code": "PROJECT_DUPLICATE", "message": "Project code or supervisor assignment already exists."}) from exc
    return dict(db.execute(text("SELECT id, code, title, status, semester_id FROM projects WHERE id = :id"), {"id": project_id}).mappings().one())


@router.get("/projects/{project_id}", response_model=ProjectDetailResponse)
def get_project_detail(project_id: int, db: Db, user: User) -> dict[str, Any]:
    _require(user, "ADMIN", "MANAGER")
    row = db.execute(
        text(
            "SELECT p.id, p.code, p.title, p.status, p.semester_id, s.code AS semester_code, m.code AS major_code "
            "FROM projects p JOIN semesters s ON s.id = p.semester_id JOIN majors m ON m.id = p.major_id WHERE p.id = :id"
        ),
        {"id": project_id},
    ).mappings().one_or_none()
    if row is None:
        raise HTTPException(status_code=404, detail={"code": "PROJECT_NOT_FOUND", "message": "Project does not exist."})
    supervisors = db.execute(
        text("SELECT l.id, l.lecturer_code, a.display_name, ps.supervisor_type FROM project_supervisors ps JOIN lecturers l ON l.id = ps.lecturer_id JOIN accounts a ON a.id = l.account_id WHERE ps.project_id = :id ORDER BY ps.supervisor_type"),
        {"id": project_id},
    ).mappings().all()
    group = db.execute(text("SELECT id, code, status FROM groups WHERE project_id = :id"), {"id": project_id}).mappings().one_or_none()
    # Keep the detail payload consistent with the project list contract.
    supervisor_items = []
    for item in supervisors:
        value = dict(item)
        value["type"] = value.pop("supervisor_type", None)
        value.pop("id", None)
        supervisor_items.append(value)
    return {**dict(row), "supervisors": supervisor_items, "group": dict(group) if group else None}


@router.patch("/groups/{group_id}", response_model=GroupResponse)
def update_group(group_id: int, payload: GroupUpdate, db: Db, user: User) -> dict[str, Any]:
    _require(user, "ADMIN", "MANAGER")
    changes = payload.model_dump(exclude_unset=True)
    if "project_id" in changes and changes["project_id"] is not None:
        project_exists = db.execute(
            text("SELECT 1 FROM projects WHERE id = :project_id"), {"project_id": changes["project_id"]}
        ).scalar_one_or_none()
        if project_exists is None:
            raise HTTPException(status_code=404, detail={"code": "PROJECT_NOT_FOUND", "message": "Project does not exist."})
        db.rollback()
    if not changes:
        row = db.execute(text("SELECT id, code, status, project_id FROM groups WHERE id = :id"), {"id": group_id}).mappings().one_or_none()
    else:
        params: dict[str, Any] = {"id": group_id}
        assignments = []
        if "code" in changes:
            params["code"] = changes["code"].strip()
            assignments.append("code = :code")
        if "project_id" in changes:
            params["project_id"] = changes["project_id"]
            assignments.append("project_id = :project_id")
        try:
            with db.begin():
                current_project = db.execute(
                    text("SELECT semester_id FROM groups g JOIN projects p ON p.id = g.project_id WHERE g.id = :id FOR UPDATE"),
                    {"id": group_id},
                ).scalar_one_or_none()
                if current_project is not None:
                    ensure_semester_writable(db, int(current_project))
                if changes.get("project_id") is not None:
                    target_semester = db.execute(
                        text("SELECT semester_id FROM projects WHERE id = :project_id FOR UPDATE"),
                        {"project_id": changes["project_id"]},
                    ).scalar_one_or_none()
                    if target_semester is not None:
                        ensure_semester_writable(db, int(target_semester))
                row = db.execute(
                    text(f"UPDATE groups SET {', '.join(assignments)} WHERE id = :id RETURNING id, code, status, project_id"),
                    params,
                ).mappings().one_or_none()
        except IntegrityError as exc:
            db.rollback()
            constraint = str(getattr(exc, "orig", exc))
            if "project_id" in changes and "project_id" in constraint:
                raise HTTPException(status_code=409, detail={"code": "PROJECT_ALREADY_ASSIGNED", "message": "Project already has a group assigned."}) from exc
            raise HTTPException(status_code=409, detail={"code": "GROUP_DUPLICATE", "message": "Group code already exists."}) from exc
    if row is None:
        raise HTTPException(status_code=404, detail={"code": "GROUP_NOT_FOUND", "message": "Group does not exist."})
    return dict(row)


@router.get("/groups/{group_id}", response_model=GroupDetailResponse)
def get_group_detail(group_id: int, db: Db, user: User) -> dict[str, Any]:
    _require(user, "ADMIN", "MANAGER")
    row = db.execute(
        text(
            "SELECT g.id, g.code, g.status, g.project_id, p.code AS project_code, p.title "
            "FROM groups g LEFT JOIN projects p ON p.id = g.project_id WHERE g.id = :id"
        ),
        {"id": group_id},
    ).mappings().one_or_none()
    if row is None:
        raise HTTPException(status_code=404, detail={"code": "GROUP_NOT_FOUND", "message": "Group does not exist."})
    members = db.execute(
        text("SELECT st.id AS student_id, st.student_code, a.display_name, a.email, gm.membership_role AS role, gm.status FROM group_memberships gm JOIN students st ON st.id = gm.student_id LEFT JOIN accounts a ON a.id = st.account_id WHERE gm.group_id = :id ORDER BY gm.membership_role DESC, st.student_code"),
        {"id": group_id},
    ).mappings().all()
    return {**dict(row), "members": [dict(item) for item in members]}


@router.get("/rounds/{round_id}/invitations", response_model=list[InvitationResponse])
def list_round_invitations(round_id: int, db: Db, user: User) -> list[dict[str, Any]]:
    _require(user, "ADMIN", "MANAGER")
    rows = db.execute(
        text(
            "SELECT ri.round_id, ri.lecturer_id, l.lecturer_code, a.display_name, a.email, "
            "CASE WHEN ri.status = 'DECLINED' THEN 'REJECTED' ELSE ri.status::text END AS status, "
            "ri.response_reason, ri.responded_at, COUNT(la.timeslot_id) FILTER (WHERE la.state = 'AVAILABLE') AS available_slot_count, "
            "MAX(la.load_preference) AS load_preference "
            "FROM round_invitations ri JOIN lecturers l ON l.id = ri.lecturer_id JOIN accounts a ON a.id = l.account_id "
            "LEFT JOIN lecturer_availabilities la ON la.round_id = ri.round_id AND la.lecturer_id = ri.lecturer_id "
            "WHERE ri.round_id = :round_id GROUP BY ri.round_id, ri.lecturer_id, l.lecturer_code, a.display_name, a.email, ri.status, ri.response_reason, ri.responded_at "
            "ORDER BY l.lecturer_code"
        ),
        {"round_id": round_id},
    ).mappings().all()
    return [dict(row) for row in rows]


@router.post("/rounds/{round_id}/invitations/{lecturer_id}/resend", response_model=ActionResponse, response_model_exclude_none=True)
def resend_invitation(round_id: int, lecturer_id: int, db: Db, user: User) -> dict[str, Any]:
    _require(user, "ADMIN", "MANAGER")
    with db.begin():
        ensure_round_semester_writable(db, round_id)
        exists = db.execute(text("SELECT 1 FROM round_invitations WHERE round_id = :round_id AND lecturer_id = :lecturer_id"), {"round_id": round_id, "lecturer_id": lecturer_id}).scalar_one_or_none()
        if exists is None:
            raise HTTPException(status_code=404, detail={"code": "INVITATION_NOT_FOUND", "message": "Invitation does not exist."})
        db.execute(text("UPDATE round_invitations SET status = 'PENDING'::invitation_status, response_reason = NULL, responded_at = NULL WHERE round_id = :round_id AND lecturer_id = :lecturer_id"), {"round_id": round_id, "lecturer_id": lecturer_id})
    return {"round_id": round_id, "lecturer_id": lecturer_id, "status": "PENDING", "resent": True}


@router.get("/rounds/{round_id}/groups", response_model=list[RoundGroupResponse])
def list_round_groups(round_id: int, db: Db, user: User) -> list[dict[str, Any]]:
    _require(user, "ADMIN", "MANAGER")
    rows = db.execute(
        text(
            "SELECT g.id AS group_id, g.code AS group_code, g.status, p.code AS project_code, p.title, "
            "COUNT(gm.id) FILTER (WHERE gm.status = 'ACTIVE') AS active_member_count, "
            "MAX(a.display_name) FILTER (WHERE gm.membership_role = 'LEADER' AND gm.status = 'ACTIVE') AS leader_name, "
            "COUNT(DISTINCT gsp.timeslot_id) FILTER (WHERE gsp.selected) AS selected_slot_count "
            "FROM round_groups rg JOIN groups g ON g.id = rg.group_id JOIN projects p ON p.id = g.project_id "
            "LEFT JOIN group_memberships gm ON gm.group_id = g.id LEFT JOIN students st ON st.id = gm.student_id "
            "LEFT JOIN accounts a ON a.id = st.account_id LEFT JOIN group_slot_preferences gsp ON gsp.round_id = rg.round_id AND gsp.group_id = g.id "
            "WHERE rg.round_id = :round_id GROUP BY g.id, g.code, g.status, p.code, p.title ORDER BY g.code"
        ),
        {"round_id": round_id},
    ).mappings().all()
    status_alias = {"PENDING_D11": "ACTIVE", "ELIGIBLE_D12": "ACTIVE", "D12_CONDITIONAL": "WAITING", "PENDING_D2": "WAITING", "COMPLETED": "COMPLETED", "FAILED": "FAILED", "DROPPED": "DROPPED"}
    return [{**dict(row), "ui_status": status_alias.get(str(row["status"]), row["status"])} for row in rows]


@router.patch("/timeslots/{timeslot_id}", response_model=TimeslotResponse)
def update_timeslot(timeslot_id: int, payload: TimeslotUpdate, db: Db, user: User) -> dict[str, Any]:
    _require(user, "ADMIN", "MANAGER")
    values = payload.model_dump(exclude_unset=True)
    if values.get("start_at") and values.get("end_at") and values["end_at"] <= values["start_at"]:
        raise HTTPException(status_code=422, detail={"code": "TIMESLOT_INVALID", "message": "end_at must be after start_at."})
    if not values:
        values = {}
    with db.begin():
        row = db.execute(text("SELECT ts.id, ts.start_at, ts.end_at, r.id AS round_id, r.status FROM timeslots ts JOIN round_days rd ON rd.id = ts.round_day_id JOIN rounds r ON r.id = rd.round_id WHERE ts.id = :id FOR UPDATE"), {"id": timeslot_id}).mappings().one_or_none()
        if row is None:
            raise HTTPException(status_code=404, detail={"code": "TIMESLOT_NOT_FOUND", "message": "Timeslot does not exist."})
        ensure_round_semester_writable(db, int(row["round_id"]))
        if str(row["status"]) not in {"DRAFT", "OPEN_REGISTRATION"}:
            raise HTTPException(status_code=409, detail={"code": "TIMESLOT_LOCKED", "message": "Timeslot cannot be changed after scheduling starts."})
        merged_start = values.get("start_at", row["start_at"])
        merged_end = values.get("end_at", row["end_at"])
        if merged_end <= merged_start:
            raise HTTPException(status_code=422, detail={"code": "TIMESLOT_INVALID", "message": "end_at must be after start_at."})
        if values:
            assignments = ", ".join(f"{key} = :{key}" for key in values)
            values["id"] = timeslot_id
            db.execute(text(f"UPDATE timeslots SET {assignments} WHERE id = :id"), values)
    return dict(db.execute(text("SELECT id, round_day_id, start_at, end_at, part, active FROM timeslots WHERE id = :id"), {"id": timeslot_id}).mappings().one())


@router.delete("/timeslots/{timeslot_id}", response_model=ActionResponse, response_model_exclude_none=True)
def disable_timeslot(timeslot_id: int, db: Db, user: User) -> dict[str, Any]:
    _require(user, "ADMIN", "MANAGER")
    with db.begin():
        row = db.execute(text("SELECT ts.id, r.id AS round_id, r.status FROM timeslots ts JOIN round_days rd ON rd.id = ts.round_day_id JOIN rounds r ON r.id = rd.round_id WHERE ts.id = :id FOR UPDATE"), {"id": timeslot_id}).mappings().one_or_none()
        if row is not None:
            ensure_round_semester_writable(db, int(row["round_id"]))
        if row is not None and str(row["status"]) not in {"DRAFT", "OPEN_REGISTRATION"}:
            raise HTTPException(status_code=409, detail={"code": "TIMESLOT_LOCKED", "message": "Timeslot cannot be changed after scheduling starts."})
        row = db.execute(text("UPDATE timeslots SET active = FALSE WHERE id = :id RETURNING id, active"), {"id": timeslot_id}).mappings().one_or_none()
    if row is None:
        raise HTTPException(status_code=404, detail={"code": "TIMESLOT_NOT_FOUND", "message": "Timeslot does not exist."})
    return dict(row)


@router.get("/semesters/{semester_id}/lecturer-quotas", response_model=list[QuotaResponse])
def list_lecturer_quotas(semester_id: int, db: Db, user: User) -> list[dict[str, Any]]:
    _require(user, "ADMIN", "MANAGER")
    rows = db.execute(
        text(
            "SELECT l.id AS lecturer_id, l.lecturer_code, a.display_name, COALESCE(q.quota, 0) AS quota, "
            "COUNT(ss.id) AS used "
            "FROM lecturers l JOIN accounts a ON a.id = l.account_id LEFT JOIN semester_lecturer_quotas q ON q.lecturer_id = l.id AND q.semester_id = :semester_id "
            "LEFT JOIN rounds r ON r.semester_id = :semester_id LEFT JOIN schedule_versions sv ON sv.round_id = r.id AND sv.status IN ('ACTIVE', 'PUBLISHED') "
            "LEFT JOIN sessions ss ON ss.schedule_version_id = sv.id "
            "LEFT JOIN council_members cm ON cm.council_id = ss.council_id AND cm.lecturer_id = l.id "
            "GROUP BY l.id, l.lecturer_code, a.display_name, q.quota ORDER BY l.lecturer_code"
        ),
        {"semester_id": semester_id},
    ).mappings().all()
    return [dict(row) for row in rows]


@router.put("/semesters/{semester_id}/lecturer-quotas/{lecturer_id}", response_model=QuotaResponse)
def set_lecturer_quota(semester_id: int, lecturer_id: int, payload: QuotaUpdate, db: Db, user: User) -> dict[str, Any]:
    _require(user, "ADMIN", "MANAGER")
    with db.begin():
        ensure_semester_writable(db, semester_id)
        row = db.execute(
            text(
                "INSERT INTO semester_lecturer_quotas (semester_id, lecturer_id, quota, updated_by) VALUES (:semester_id, :lecturer_id, :quota, :updated_by) "
                "ON CONFLICT (semester_id, lecturer_id) DO UPDATE SET quota = EXCLUDED.quota, updated_by = EXCLUDED.updated_by, updated_at = now() "
                "RETURNING semester_id, lecturer_id, quota, updated_at"
            ),
            {"semester_id": semester_id, "lecturer_id": lecturer_id, "quota": payload.quota, "updated_by": user.account_id},
        ).mappings().one()
    return dict(row)


@router.get("/sessions", response_model=list[SessionResponse])
def list_sessions(db: Db, user: User, round_id: int | None = None, version_id: int | None = None, status_filter: str | None = None) -> list[dict[str, Any]]:
    _require(user, "ADMIN", "MANAGER")
    rows = db.execute(
        text(
            "SELECT s.id, sa.id AS assignment_id, s.schedule_version_id AS version_id, sv.round_id, s.group_id, g.code AS group_code, p.code AS project_code, "
            "s.timeslot_id, s.room_id, rm.code AS room_code, s.start_at, s.end_at, s.status, "
            "jsonb_agg(DISTINCT jsonb_build_object('id', l.id, 'code', l.lecturer_code, 'name', a.display_name)) FILTER (WHERE l.id IS NOT NULL) AS reviewers "
            "FROM sessions s JOIN schedule_versions sv ON sv.id = s.schedule_version_id JOIN groups g ON g.id = s.group_id JOIN schedule_assignments sa ON sa.schedule_version_id=s.schedule_version_id AND sa.group_id=s.group_id JOIN projects p ON p.id = sa.project_id "
            "LEFT JOIN rooms rm ON rm.id = s.room_id LEFT JOIN council_members cm ON cm.council_id = s.council_id LEFT JOIN lecturers l ON l.id = cm.lecturer_id LEFT JOIN accounts a ON a.id = l.account_id "
            "WHERE (CAST(:round_id AS BIGINT) IS NULL OR sv.round_id = CAST(:round_id AS BIGINT)) AND (CAST(:version_id AS BIGINT) IS NULL OR s.schedule_version_id = CAST(:version_id AS BIGINT)) AND (CAST(:status_filter AS TEXT) IS NULL OR s.status::text = CAST(:status_filter AS TEXT)) "
            "GROUP BY s.id, sa.id, s.schedule_version_id, sv.round_id, s.group_id, g.code, p.code, "
            "s.timeslot_id, s.room_id, rm.code, s.start_at, s.end_at, s.status "
            "ORDER BY s.start_at, s.id"
        ),
        {"round_id": round_id, "version_id": version_id, "status_filter": status_filter},
    ).mappings().all()
    return [dict(row) for row in rows]


@router.get("/sessions/{session_id}", response_model=SessionResponse)
def get_session_detail(session_id: int, db: Db, user: User) -> dict[str, Any]:
    _require(user, "ADMIN", "MANAGER")
    rows = list_sessions(db, user)
    for row in rows:
        if row["id"] == session_id:
            return row
    raise HTTPException(status_code=404, detail={"code": "SESSION_NOT_FOUND", "message": "Session does not exist."})


@router.get("/reschedule-requests", response_model=list[RescheduleRequestResponse])
def list_reschedule_requests(db: Db, user: User, status_filter: str | None = None) -> list[dict[str, Any]]:
    _require(user, "ADMIN", "MANAGER")
    rows = db.execute(
        text(
            "SELECT rr.id, rr.session_id, rr.requested_by, aa.display_name AS requester_name, rr.reason, rr.status, rr.reviewed_by, rr.created_at, rr.reviewed_at, "
            "s.group_id, g.code AS group_code, s.start_at, s.end_at, rm.code AS room_code "
            "FROM reschedule_requests rr JOIN sessions s ON s.id = rr.session_id JOIN groups g ON g.id = s.group_id JOIN accounts aa ON aa.id = rr.requested_by LEFT JOIN rooms rm ON rm.id = s.room_id "
            "WHERE (CAST(:status_filter AS TEXT) IS NULL OR rr.status::text = CAST(:status_filter AS TEXT)) ORDER BY rr.created_at DESC"
        ),
        {"status_filter": status_filter},
    ).mappings().all()
    return [dict(row) for row in rows]


@router.get("/reports/group-progress")
def group_progress_report(
    semester_id: int,
    db: Db,
    user: User,
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, alias="pageSize", ge=1, le=200),
) -> dict[str, Any]:
    _require(user, "ADMIN", "MANAGER")
    offset = (page - 1) * page_size
    rows = db.execute(
        text(
            "WITH latest_results AS ("
            " SELECT s.group_id, a.project_id, pr.title AS project_name, r.type, sr.outcome::text AS outcome, sr.entered_at, sr.verifier_lecturer_id, "
            " ROW_NUMBER() OVER (PARTITION BY s.group_id, r.type ORDER BY sr.entered_at DESC, sr.id DESC) AS rn "
            " FROM session_results sr JOIN sessions s ON s.id = sr.session_id JOIN schedule_assignments a ON a.schedule_version_id=s.schedule_version_id AND a.group_id=s.group_id JOIN projects pr ON pr.id=a.project_id JOIN schedule_versions sv ON sv.id = s.schedule_version_id "
            " JOIN rounds r ON r.id = sv.round_id WHERE r.semester_id = :semester_id AND sv.status IN ('ACTIVE', 'PUBLISHED')"
            "), latest_remediation AS ("
            " SELECT rc.group_id, rc.status::text AS remediation_status, rc.due_at, rc.verifier_lecturer_id, "
            " ROW_NUMBER() OVER (PARTITION BY rc.group_id ORDER BY rc.due_at DESC, rc.id DESC) AS rn "
            " FROM remediation_cases rc JOIN session_results rrs ON rrs.id = rc.session_result_id "
            " JOIN sessions rss ON rss.id = rrs.session_id "
            " JOIN schedule_assignments ra ON ra.schedule_version_id = rss.schedule_version_id AND ra.group_id = rss.group_id "
            " JOIN projects rp ON rp.id = ra.project_id WHERE rp.semester_id = :semester_id"
            "), historical_groups AS ("
            " SELECT DISTINCT ON (a.group_id) a.group_id, a.project_id, p.title AS project_name"
            " FROM schedule_assignments a JOIN projects p ON p.id = a.project_id"
            " JOIN schedule_versions hv ON hv.id = a.schedule_version_id"
            " WHERE p.semester_id = :semester_id AND hv.status IN ('ACTIVE', 'PUBLISHED')"
            " ORDER BY a.group_id, hv.activated_at DESC NULLS LAST, a.id DESC"
            ") SELECT hg.group_id AS group_id, g.code AS group_code, COALESCE(MAX(lr.project_name), hg.project_name) AS project_name, g.status AS group_status, "
            "MAX(lr.outcome) FILTER (WHERE lr.type = 'REVIEW_1' AND lr.rn = 1) AS review_1, "
            "MAX(lr.outcome) FILTER (WHERE lr.type = 'REVIEW_2' AND lr.rn = 1) AS review_2, "
            "MAX(lr.outcome) FILTER (WHERE lr.type = 'DEFENSE_1_1' AND lr.rn = 1) AS defense_1_1, "
            "MAX(lr.outcome) FILTER (WHERE lr.type = 'DEFENSE_1_2' AND lr.rn = 1) AS defense_1_2, "
            "MAX(lr.outcome) FILTER (WHERE lr.type = 'DEFENSE_2' AND lr.rn = 1) AS defense_2, "
            "MAX(lr.verifier_lecturer_id) FILTER (WHERE lr.type = 'DEFENSE_1_1' AND lr.rn = 1) AS result_verifier_lecturer_id, "
            "MAX(lm.remediation_status) FILTER (WHERE lm.rn = 1) AS remediation_status, "
            "MAX(lm.due_at) FILTER (WHERE lm.rn = 1) AS remediation_due_at, "
            "MAX(lm.verifier_lecturer_id) FILTER (WHERE lm.rn = 1) AS remediation_verifier_lecturer_id, "
            "COUNT(*) OVER() AS total_count "
            "FROM historical_groups hg JOIN groups g ON g.id = hg.group_id LEFT JOIN latest_results lr ON lr.group_id = hg.group_id "
            "LEFT JOIN latest_remediation lm ON lm.group_id = hg.group_id "
            "GROUP BY hg.group_id, g.code, hg.project_name, g.status ORDER BY g.code "
            "LIMIT :limit OFFSET :offset"
        ),
        {"semester_id": semester_id, "limit": page_size, "offset": offset},
    ).mappings().all()
    total = rows[0]["total_count"] if rows else 0
    items = [{k: v for k, v in dict(row).items() if k != "total_count"} for row in rows]
    return success_payload(items, meta={"page": page, "pageSize": page_size, "total": total})


@router.get("/results", response_model=list[ResultResponse])
def list_results(db: Db, user: User, round_id: int | None = None) -> list[dict[str, Any]]:
    _require(user, "ADMIN", "MANAGER")
    rows = db.execute(
        text(
            "SELECT sr.id, sr.session_id, s.group_id, g.code AS group_code, r.type AS round_type, sr.outcome, sr.note, sr.entered_at, sr.verify_status, sr.remediation_due_at, sr.verifier_lecturer_id "
            "FROM session_results sr JOIN sessions s ON s.id = sr.session_id JOIN groups g ON g.id = s.group_id JOIN schedule_versions sv ON sv.id = s.schedule_version_id JOIN rounds r ON r.id = sv.round_id WHERE (CAST(:round_id AS BIGINT) IS NULL OR r.id = CAST(:round_id AS BIGINT)) ORDER BY sr.entered_at DESC"
        ),
        {"round_id": round_id},
    ).mappings().all()
    return [dict(row) for row in rows]


def _normalise_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").strip().lower())


def _workbook_rows(upload: UploadFile) -> dict[str, list[dict[str, Any]]]:
    try:
        workbook = load_workbook(upload.file, read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=422, detail={"code": "IMPORT_INVALID_FILE", "message": "Only a readable .xlsx file is supported."}) from exc
    result: dict[str, list[dict[str, Any]]] = {}
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [_normalise_header(item) for item in rows[0]]
        result[_normalise_header(sheet.title)] = [
            {headers[index]: row[index] for index in range(min(len(headers), len(row))) if headers[index]}
            for row in rows[1:] if any(value is not None for value in row)
        ]
    return result


_LECTURER_IMPORT_MAX_FILE_BYTES = 5 * 1024 * 1024
_LECTURER_IMPORT_MAX_ROWS = 2000


def _lecturer_import_rows(upload: UploadFile) -> list[tuple[int, dict[str, Any]]]:
    raw = upload.file.read(_LECTURER_IMPORT_MAX_FILE_BYTES + 1)
    if len(raw) > _LECTURER_IMPORT_MAX_FILE_BYTES:
        raise HTTPException(status_code=422, detail={"code": "IMPORT_FILE_TOO_LARGE", "message": "File exceeds the 5MB import limit."})
    try:
        workbook = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=422, detail={"code": "IMPORT_INVALID_FILE", "message": "Only a readable .xlsx file is supported."}) from exc
    sheet = workbook.worksheets[0]
    header_row: int | None = None
    result: list[tuple[int, dict[str, Any]]] = []
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if header_row is None:
            if row and _normalise_header(row[0]) == "stt" and len(row) > 3 and "email" in _normalise_header(row[3]):
                header_row = row_number
            continue
        if not row or not any(cell is not None for cell in row):
            continue
        if len(result) >= _LECTURER_IMPORT_MAX_ROWS:
            raise HTTPException(status_code=422, detail={"code": "IMPORT_TOO_MANY_ROWS", "message": f"Import is limited to {_LECTURER_IMPORT_MAX_ROWS} data rows."})
        result.append((row_number, {
            "lecturer_code": row[1] if len(row) > 1 else None,
            "display_name": row[2] if len(row) > 2 else None,
            "email": row[3] if len(row) > 3 else None,
        }))
    if header_row is None:
        raise HTTPException(status_code=422, detail={"code": "IMPORT_INVALID_FILE", "message": "Expected the lecturers_template.xlsx header row (STT | Mã giảng viên | Họ và tên | Email)."})
    return result


@router.post("/lecturers/import", status_code=status.HTTP_201_CREATED, response_model=LecturerImportResponse)
async def import_lecturers(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: CurrentUser = Depends(get_current_user),
) -> dict[str, Any]:
    _require(user, "ADMIN", "MANAGER")
    rows = _lecturer_import_rows(file)
    created = 0
    errors: list[dict[str, Any]] = []
    accounts: list[dict[str, Any]] = []
    with db.begin():
        for row_number, row in rows:
            code = str(row.get("lecturer_code") or "").strip().upper()
            display_name = str(row.get("display_name") or "").strip()
            email = str(row.get("email") or "").strip().lower()
            if not code or not display_name or not email:
                errors.append({"row": row_number, "code": "REQUIRED_FIELD_MISSING", "message": "lecturer_code, display_name and email are all required."})
                continue
            if len(code) > 32 or len(display_name) > 160 or len(email) > 320 or "@" not in email:
                errors.append({"row": row_number, "code": "FIELD_INVALID", "message": "lecturer_code/display_name/email exceed the allowed length or email is malformed."})
                continue
            temp_password = secrets.token_urlsafe(9)
            try:
                with db.begin_nested():
                    account_id = db.execute(
                        text("INSERT INTO accounts (email, display_name, password_hash) VALUES (:email, :display_name, :password_hash) RETURNING id"),
                        {"email": email, "display_name": display_name, "password_hash": password_hasher.hash(temp_password)},
                    ).scalar_one()
                    db.execute(text("INSERT INTO account_roles (account_id, role) VALUES (:account_id, 'LECTURER')"), {"account_id": account_id})
                    lecturer_id = db.execute(
                        text("INSERT INTO lecturers (account_id, lecturer_code) VALUES (:account_id, :code) RETURNING id"),
                        {"account_id": account_id, "code": normalize_code(code)},
                    ).scalar_one()
            except IntegrityError:
                errors.append({"row": row_number, "code": "LECTURER_DUPLICATE", "message": "Email or lecturer code already exists."})
                continue
            created += 1
            accounts.append({"row": row_number, "lecturer_id": lecturer_id, "lecturer_code": code, "email": email, "display_name": display_name, "temp_password": temp_password})
        db.execute(
            text("INSERT INTO audit_events (actor_id, action, entity_type, entity_id, after_json) VALUES (:actor_id, 'LECTURERS_IMPORTED', 'lecturer', :entity_id, CAST(:after_json AS JSONB))"),
            {"actor_id": _actor_id(db, user), "entity_id": "bulk", "after_json": _json({"created": created, "skipped": len(errors)})},
        )
    return {"created": created, "skipped": len(errors), "errors": errors, "accounts": accounts}


@router.post("/projects/import", status_code=status.HTTP_201_CREATED, response_model=ImportResponse)
async def import_projects(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: CurrentUser = Depends(get_current_user),
) -> dict[str, Any]:
    _require(user, "ADMIN", "MANAGER")
    sheets = _workbook_rows(file)
    rows = sheets.get("projects") or next(iter(sheets.values()), [])
    created = 0
    errors: list[dict[str, Any]] = []
    with db.begin():
        for index, row in enumerate(rows, start=2):
            code = str(row.get("code") or row.get("projectcode") or "").strip()
            title = str(row.get("title") or row.get("project") or row.get("name") or code).strip()
            semester_code = str(row.get("semestercode") or row.get("semester") or "").strip()
            major_code = str(row.get("majorcode") or row.get("major") or "").strip()
            if not code or not semester_code or not major_code:
                errors.append({"row": index, "code": "REQUIRED_FIELD_MISSING"})
                continue
            semester_id = db.execute(text("SELECT id FROM semesters WHERE code = :code"), {"code": semester_code}).scalar_one_or_none()
            major_id = db.execute(text("SELECT id FROM majors WHERE code = :code"), {"code": major_code}).scalar_one_or_none()
            if semester_id is None or major_id is None:
                errors.append({"row": index, "code": "SEMESTER_OR_MAJOR_NOT_FOUND"})
                continue
            ensure_semester_writable(db, int(semester_id))
            try:
                with db.begin_nested():
                    db.execute(text("INSERT INTO projects (semester_id, major_id, code, title) VALUES (:semester_id, :major_id, :code, :title)"), {"semester_id": semester_id, "major_id": major_id, "code": code, "title": title})
                created += 1
            except Exception:
                errors.append({"row": index, "code": "PROJECT_DUPLICATE_OR_INVALID"})
    return {"created": created, "skipped": len(errors), "errors": errors}


@router.post("/groups/import", status_code=status.HTTP_201_CREATED, response_model=ImportResponse)
async def import_groups(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: CurrentUser = Depends(get_current_user),
) -> dict[str, Any]:
    _require(user, "ADMIN", "MANAGER")
    sheets = _workbook_rows(file)
    rows = sheets.get("groups") or next(iter(sheets.values()), [])
    grouped: dict[tuple[str, str, str], list[dict[str, Any]]] = {}
    for row in rows:
        project = str(row.get("projectcode") or row.get("project") or "").strip()
        group = str(row.get("groupcode") or row.get("group") or row.get("code") or "").strip()
        semester = str(row.get("semestercode") or row.get("semester") or "").strip()
        if project and group:
            grouped.setdefault((project, group, semester), []).append(row)
    created = 0
    errors: list[dict[str, Any]] = []
    with db.begin():
        for (project_code, group_code, semester_code), members in grouped.items():
            project_query = "SELECT p.id FROM projects p JOIN semesters s ON s.id = p.semester_id WHERE p.code = :code"
            project_params: dict[str, Any] = {"code": project_code}
            if semester_code:
                project_query += " AND s.code = :semester_code"
                project_params["semester_code"] = semester_code
            project_ids = db.execute(text(project_query), project_params).scalars().all()
            project_id = project_ids[0] if len(project_ids) == 1 else None
            if project_id is None:
                errors.append({"group_code": group_code, "code": "PROJECT_NOT_FOUND_OR_AMBIGUOUS"})
                continue
            semester_id = db.execute(text("SELECT semester_id FROM projects WHERE id = :project_id"), {"project_id": project_id}).scalar_one()
            ensure_semester_writable(db, int(semester_id))
            try:
                with db.begin_nested():
                    group_id = db.execute(text("INSERT INTO groups (project_id, code) VALUES (:project_id, :code) RETURNING id"), {"project_id": project_id, "code": group_code}).scalar_one()
                    leader_count = 0
                    for member in members:
                        student_code = str(member.get("studentcode") or member.get("student") or member.get("studentid") or "").strip()
                        if not student_code:
                            continue
                        student_id = db.execute(text("SELECT id FROM students WHERE student_code = :code"), {"code": student_code}).scalar_one_or_none()
                        if student_id is None:
                            raise ValueError(f"Student {student_code} not found")
                        role = "LEADER" if str(member.get("role") or "MEMBER").upper() == "LEADER" else "MEMBER"
                        leader_count += role == "LEADER"
                        db.execute(text("INSERT INTO group_memberships (group_id, student_id, membership_role) VALUES (:group_id, :student_id, CAST(:role AS membership_role))"), {"group_id": group_id, "student_id": student_id, "role": role})
                    if leader_count != 1:
                        raise ValueError("Exactly one leader is required")
                created += 1
            except Exception:
                errors.append({"group_code": group_code, "code": "GROUP_INVALID", "message": "Invalid group row; verify project, students and exactly one leader."})
    return {"created": created, "skipped": len(errors), "errors": errors}


def _xlsx_response(workbook: Workbook, filename: str) -> StreamingResponse:
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@router.get("/exports/round/{round_id}.xlsx")
def export_round(round_id: int, db: Db, user: User) -> StreamingResponse:
    _require(user, "ADMIN", "MANAGER")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Schedule"
    sheet.append(["Session", "Group", "Round", "Start", "End", "Room", "Status"])
    rows = db.execute(text("SELECT s.id, g.code AS group_code, r.type, s.start_at, s.end_at, rm.code AS room_code, s.status FROM sessions s JOIN schedule_versions sv ON sv.id = s.schedule_version_id JOIN rounds r ON r.id = sv.round_id JOIN groups g ON g.id = s.group_id LEFT JOIN rooms rm ON rm.id = s.room_id WHERE r.id = :round_id AND sv.activated_at IS NOT NULL AND sv.status IN ('ACTIVE', 'PUBLISHED') ORDER BY s.start_at"), {"round_id": round_id}).mappings().all()
    for row in rows:
        sheet.append([row[key] for key in ("id", "group_code", "type", "start_at", "end_at", "room_code", "status")])
    return _xlsx_response(workbook, f"round-{round_id}.xlsx")


@router.get("/exports/semester/{semester_id}/schedule.xlsx")
def export_semester_schedule(semester_id: int, db: Db, user: User) -> StreamingResponse:
    _require(user, "ADMIN", "MANAGER")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Semester Schedule"
    sheet.append(["Session", "Group", "Round", "Start", "End", "Room", "Status"])
    rows = db.execute(text("SELECT s.id, g.code AS group_code, r.type, s.start_at, s.end_at, rm.code AS room_code, s.status FROM sessions s JOIN schedule_versions sv ON sv.id = s.schedule_version_id JOIN rounds r ON r.id = sv.round_id JOIN groups g ON g.id = s.group_id LEFT JOIN rooms rm ON rm.id = s.room_id WHERE r.semester_id = :semester_id AND sv.activated_at IS NOT NULL AND sv.status IN ('ACTIVE', 'PUBLISHED') ORDER BY s.start_at"), {"semester_id": semester_id}).mappings().all()
    for row in rows:
        sheet.append([row[key] for key in ("id", "group_code", "type", "start_at", "end_at", "room_code", "status")])
    return _xlsx_response(workbook, f"semester-{semester_id}-schedule.xlsx")


@router.get("/exports/semester/{semester_id}/results.xlsx")
def export_semester_results(semester_id: int, db: Db, user: User) -> StreamingResponse:
    _require(user, "ADMIN", "MANAGER")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Results"
    sheet.append(["Group", "Round", "Outcome", "Entered At", "Verify Status"])
    rows = db.execute(text("SELECT g.code AS group_code, r.type, sr.outcome, sr.entered_at, sr.verify_status FROM session_results sr JOIN sessions s ON s.id = sr.session_id JOIN schedule_versions sv ON sv.id = s.schedule_version_id JOIN rounds r ON r.id = sv.round_id JOIN groups g ON g.id = s.group_id WHERE r.semester_id = :semester_id AND sv.activated_at IS NOT NULL AND sv.status IN ('ACTIVE', 'PUBLISHED') ORDER BY sr.entered_at"), {"semester_id": semester_id}).mappings().all()
    for row in rows:
        sheet.append([row[key] for key in ("group_code", "type", "outcome", "entered_at", "verify_status")])
    return _xlsx_response(workbook, f"semester-{semester_id}-results.xlsx")
