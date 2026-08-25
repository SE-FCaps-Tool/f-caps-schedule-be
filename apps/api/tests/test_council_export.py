"""Integration tests for the Defense-council xlsx export."""

from __future__ import annotations

from datetime import datetime, timedelta
from io import BytesIO
from uuid import uuid4

import pytest
from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.config import get_settings
from app.database import get_engine
from app.services.councils import create_council

HEADERS = {"X-Test-Session": "active-manager"}


@pytest.fixture()
def lecturer_ids():
    engine = get_engine(get_settings().database_url)
    marker = uuid4().hex[:8]
    ids: list[int] = []
    with Session(engine) as db, db.begin():
        for index in range(10):
            account_id = db.execute(
                text(
                    "INSERT INTO accounts (email, display_name, password_hash) "
                    "VALUES (:email, :name, 'x') RETURNING id"
                ),
                {"email": f"council-export-{marker}-{index}@example.com", "name": f"Council Export Lecturer {index}"},
            ).scalar_one()
            db.execute(
                text("INSERT INTO account_roles (account_id, role) VALUES (:account_id, 'LECTURER')"),
                {"account_id": account_id},
            )
            lecturer_id = db.execute(
                text("INSERT INTO lecturers (account_id, lecturer_code) VALUES (:account_id, :code) RETURNING id"),
                {"account_id": account_id, "code": f"CEXP{marker}{index}"},
            ).scalar_one()
            ids.append(int(lecturer_id))
    yield ids
    # Lecturers that ended up on a sealed (immutable) Council can never be
    # deleted; best-effort cleanup only, matching the other Council tests.
    try:
        with Session(engine) as db, db.begin():
            account_ids = db.execute(
                text("SELECT account_id FROM lecturers WHERE id = ANY(:ids)"), {"ids": ids}
            ).scalars().all()
            db.execute(text("DELETE FROM lecturers WHERE id = ANY(:ids)"), {"ids": ids})
            db.execute(text("DELETE FROM account_roles WHERE account_id = ANY(:ids)"), {"ids": list(account_ids)})
            db.execute(text("DELETE FROM accounts WHERE id = ANY(:ids)"), {"ids": list(account_ids)})
    except IntegrityError:
        pass


@pytest.fixture()
def round_id():
    engine = get_engine(get_settings().database_url)
    with Session(engine) as db, db.begin():
        semester_id = db.execute(text("SELECT id FROM semesters ORDER BY id LIMIT 1")).scalar_one()
        new_id = db.execute(
            text(
                "INSERT INTO rounds "
                "(semester_id, name, type, status, session_duration_minutes, reviewer_count) "
                "VALUES (:semester_id, 'Council Export Test', CAST('DEFENSE_1_2' AS round_type), "
                "CAST('DRAFT' AS round_status), 60, 5) RETURNING id"
            ),
            {"semester_id": semester_id},
        ).scalar_one()
    yield int(new_id)
    # A Round that grew a sealed Council can never be deleted (RESTRICT FK);
    # best-effort cleanup only, matching the other Council tests.
    try:
        with Session(engine) as db, db.begin():
            db.execute(text("DELETE FROM rounds WHERE id = :id"), {"id": new_id})
    except IntegrityError:
        pass


def _make_group(db: Session, semester_id: int, major_id: int, marker: str, suffix: str) -> tuple[int, int]:
    project_id = db.execute(
        text(
            "INSERT INTO projects (semester_id, major_id, code, title) "
            "VALUES (:semester_id, :major_id, :code, :title) RETURNING id"
        ),
        {
            "semester_id": semester_id,
            "major_id": major_id,
            "code": f"CEXP-{marker}-{suffix}",
            "title": f"Council export project {suffix}",
        },
    ).scalar_one()
    group_id = db.execute(
        text("INSERT INTO groups (project_id, code) VALUES (:project_id, :code) RETURNING id"),
        {"project_id": project_id, "code": f"CEXP-G-{marker}-{suffix}"},
    ).scalar_one()
    return int(project_id), int(group_id)


def _make_session(
    db: Session,
    *,
    round_id: int,
    version_id: int,
    project_id: int,
    group_id: int,
    council_id: int,
    day_date: str,
) -> None:
    day_id = db.execute(
        text("INSERT INTO round_days (round_id, day_date) VALUES (:round_id, :day_date) RETURNING id"),
        {"round_id": round_id, "day_date": day_date},
    ).scalar_one()
    start_at = datetime.fromisoformat(f"{day_date}T01:00:00+00:00")
    end_at = start_at + timedelta(hours=1)
    timeslot_id = db.execute(
        text(
            "INSERT INTO timeslots (round_day_id, start_at, end_at) "
            "VALUES (:day_id, :start_at, :end_at) RETURNING id"
        ),
        {"day_id": day_id, "start_at": start_at, "end_at": end_at},
    ).scalar_one()
    db.execute(
        text(
            "INSERT INTO schedule_assignments (schedule_version_id, group_id, project_id, timeslot_id, start_at, end_at) "
            "VALUES (:version_id, :group_id, :project_id, :timeslot_id, :start_at, :end_at)"
        ),
        {
            "version_id": version_id,
            "group_id": group_id,
            "project_id": project_id,
            "timeslot_id": timeslot_id,
            "start_at": start_at,
            "end_at": end_at,
        },
    )
    db.execute(
        text(
            "INSERT INTO sessions (schedule_version_id, group_id, timeslot_id, council_id, start_at, end_at, status) "
            "VALUES (:version_id, :group_id, :timeslot_id, :council_id, :start_at, :end_at, 'SCHEDULED')"
        ),
        {
            "version_id": version_id,
            "group_id": group_id,
            "timeslot_id": timeslot_id,
            "council_id": council_id,
            "start_at": start_at,
            "end_at": end_at,
        },
    )


@pytest.mark.integration
def test_council_export_recovers_committee_seat_order_and_falls_back_to_lecturer_id(client, round_id, lecturer_ids):
    engine = get_engine(get_settings().database_url)
    marker = uuid4().hex[:8]
    committee_member_ids = lecturer_ids[:5]
    fallback_member_ids = sorted(lecturer_ids[5:])

    # Committee created with a shuffled member order: role assignment follows
    # submission order (CHAIR, SECRETARY, MEMBER x3), not ascending lecturer_id.
    shuffled = [committee_member_ids[4], committee_member_ids[0], committee_member_ids[3], committee_member_ids[1], committee_member_ids[2]]
    created = client.post(
        "/api/v1/committees",
        headers=HEADERS,
        json={"groups": [{"code": f"CEXP-CMT-{marker}", "memberIds": shuffled}]},
    )
    assert created.status_code == 201, created.text
    committee_id = created.json()["data"]["committees"][0]["id"]
    assigned = client.put(
        f"/api/v1/rounds/{round_id}/committees",
        headers=HEADERS,
        json={"committeeIds": [f"cmt_{committee_id}"]},
    )
    assert assigned.status_code == 200, assigned.text

    with Session(engine) as db, db.begin():
        semester_id = db.execute(text("SELECT semester_id FROM rounds WHERE id = :id"), {"id": round_id}).scalar_one()
        major_id = db.execute(text("SELECT id FROM majors ORDER BY id LIMIT 1")).scalar_one()
        matched_project_id, matched_group_id = _make_group(db, semester_id, major_id, marker, "matched")
        fallback_project_id, fallback_group_id = _make_group(db, semester_id, major_id, marker, "fallback")

        version_id = db.execute(
            text(
                "INSERT INTO schedule_versions (round_id, version_no, status, activated_at) "
                "VALUES (:round_id, 1, 'ACTIVE', now()) RETURNING id"
            ),
            {"round_id": round_id},
        ).scalar_one()

        names = {lecturer_id: f"Council Export Lecturer {index}" for index, lecturer_id in enumerate(lecturer_ids)}

        matched_council_id = create_council(
            db,
            round_id,
            [{"lecturer_id": lid, "snapshot_name": names[lid]} for lid in committee_member_ids],
        )
        _make_session(
            db,
            round_id=round_id,
            version_id=int(version_id),
            project_id=matched_project_id,
            group_id=matched_group_id,
            council_id=matched_council_id,
            day_date="2055-04-01",
        )

        fallback_council_id = create_council(
            db,
            round_id,
            [{"lecturer_id": lid, "snapshot_name": names[lid]} for lid in fallback_member_ids],
        )
        _make_session(
            db,
            round_id=round_id,
            version_id=int(version_id),
            project_id=fallback_project_id,
            group_id=fallback_group_id,
            council_id=fallback_council_id,
            day_date="2055-04-02",
        )

    response = client.get(f"/api/v1/exports/round/{round_id}/council.xlsx", headers=HEADERS)
    assert response.status_code == 200, response.text
    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0] == (
        "STT",
        "Mã nhóm",
        "Mã đề tài",
        "Tên đề tài",
        "Ngày bảo vệ",
        "Giờ bảo vệ",
        "Phòng",
        "Chủ tịch hội đồng",
        "Thư ký hội đồng",
        "Thành viên hội đồng",
        "Thành viên hội đồng",
        "Thành viên hội đồng",
    )
    by_group_code = {row[1]: row for row in rows[1:]}

    matched_row = by_group_code[f"CEXP-G-{marker}-matched"]
    assert matched_row[4:7] == ("01/04/2055", "08:00-09:00", None)
    assert matched_row[7:12] == (
        names[shuffled[0]],
        names[shuffled[1]],
        names[shuffled[2]],
        names[shuffled[3]],
        names[shuffled[4]],
    )

    fallback_row = by_group_code[f"CEXP-G-{marker}-fallback"]
    assert fallback_row[4:7] == ("02/04/2055", "08:00-09:00", None)
    assert fallback_row[7:12] == tuple(names[lid] for lid in fallback_member_ids)

    header_cell = sheet.cell(row=1, column=1)
    assert header_cell.font.bold is True
    assert header_cell.fill.start_color.rgb == "00305496"
    title_column_width = sheet.column_dimensions["D"].width
    stt_column_width = sheet.column_dimensions["A"].width
    assert title_column_width > stt_column_width
    assert title_column_width >= len("Council export project matched") + 2

    with Session(engine) as db, db.begin():
        db.execute(text("DELETE FROM sessions WHERE group_id = ANY(:ids)"), {"ids": [matched_group_id, fallback_group_id]})
        db.execute(text("DELETE FROM schedule_assignments WHERE group_id = ANY(:ids)"), {"ids": [matched_group_id, fallback_group_id]})
        db.execute(text("DELETE FROM schedule_versions WHERE id = :id"), {"id": version_id})
        db.execute(text("DELETE FROM groups WHERE id = ANY(:ids)"), {"ids": [matched_group_id, fallback_group_id]})
        db.execute(text("DELETE FROM projects WHERE id = ANY(:ids)"), {"ids": [matched_project_id, fallback_project_id]})
        db.execute(text("DELETE FROM round_days WHERE round_id = :id"), {"id": round_id})
        db.execute(text("DELETE FROM round_committees WHERE committee_id = :id"), {"id": committee_id})
        db.execute(text("DELETE FROM committees WHERE id = :id"), {"id": committee_id})


@pytest.mark.integration
def test_council_export_uses_plain_seats_for_reviewer_only_rounds(client, lecturer_ids):
    """A Round configured with reviewer_count <= 3 has no Chair/Secretary split (assign_roles semantics)."""
    engine = get_engine(get_settings().database_url)
    marker = uuid4().hex[:8]
    with Session(engine) as db, db.begin():
        semester_id = db.execute(text("SELECT id FROM semesters ORDER BY id LIMIT 1")).scalar_one()
        round_id = db.execute(
            text(
                "INSERT INTO rounds (semester_id, name, type, status, session_duration_minutes, reviewer_count) "
                "VALUES (:semester_id, 'Council Export 3-Seat Test', CAST('DEFENSE_1_1' AS round_type), "
                "CAST('DRAFT' AS round_status), 60, 3) RETURNING id"
            ),
            {"semester_id": semester_id},
        ).scalar_one()
        major_id = db.execute(text("SELECT id FROM majors ORDER BY id LIMIT 1")).scalar_one()
        project_id, group_id = _make_group(db, semester_id, major_id, marker, "threeseat")
        version_id = db.execute(
            text(
                "INSERT INTO schedule_versions (round_id, version_no, status, activated_at) "
                "VALUES (:round_id, 1, 'ACTIVE', now()) RETURNING id"
            ),
            {"round_id": round_id},
        ).scalar_one()
        member_ids = sorted(lecturer_ids[:3])
        names = {lid: f"Council Export Lecturer {index}" for index, lid in enumerate(lecturer_ids)}
        council_id = create_council(
            db, round_id, [{"lecturer_id": lid, "snapshot_name": names[lid]} for lid in member_ids]
        )
        _make_session(
            db,
            round_id=round_id,
            version_id=int(version_id),
            project_id=project_id,
            group_id=group_id,
            council_id=council_id,
            day_date="2055-05-01",
        )

    response = client.get(f"/api/v1/exports/round/{round_id}/council.xlsx", headers=HEADERS)
    assert response.status_code == 200, response.text
    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0] == (
        "STT",
        "Mã nhóm",
        "Mã đề tài",
        "Tên đề tài",
        "Ngày bảo vệ",
        "Giờ bảo vệ",
        "Phòng",
        "Thành viên hội đồng",
        "Thành viên hội đồng",
        "Thành viên hội đồng",
    )
    assert rows[1][7:10] == tuple(names[lid] for lid in member_ids)

    with Session(engine) as db, db.begin():
        db.execute(text("DELETE FROM sessions WHERE group_id = :id"), {"id": group_id})
        db.execute(text("DELETE FROM schedule_assignments WHERE group_id = :id"), {"id": group_id})
        db.execute(text("DELETE FROM schedule_versions WHERE id = :id"), {"id": version_id})
        db.execute(text("DELETE FROM groups WHERE id = :id"), {"id": group_id})
        db.execute(text("DELETE FROM projects WHERE id = :id"), {"id": project_id})
        db.execute(text("DELETE FROM round_days WHERE round_id = :id"), {"id": round_id})
    try:
        with Session(engine) as db, db.begin():
            db.execute(text("DELETE FROM rounds WHERE id = :id"), {"id": round_id})
    except IntegrityError:
        pass


@pytest.mark.integration
def test_council_export_rejects_unknown_round(client):
    response = client.get("/api/v1/exports/round/999999999/council.xlsx", headers=HEADERS)
    assert response.status_code == 404, response.text
    assert response.json()["error"]["code"] == "ROUND_NOT_FOUND"


@pytest.mark.integration
def test_council_export_requires_admin_or_manager(client, round_id):
    response = client.get(
        f"/api/v1/exports/round/{round_id}/council.xlsx", headers={"X-Test-Session": "active-student"}
    )
    assert response.status_code == 403, response.text
