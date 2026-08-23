"""Import Defense 1.2 eligibility from the SU26 one-sheet workbook.

The workbook marks projects that passed Review 1.1 with ``Bảo vệ kỳ 2``.
By default this importer only updates eligible groups to ``ELIGIBLE_D12``.
Pass ``--attach-round`` to also create (or reuse) a draft ``DEFENSE_1_2`` round
and link those groups. It never invents students or groups for projects that
are incomplete in the database.

Default mode is validation-only.  Pass ``--apply`` to write to PostgreSQL.
The operation is idempotent for the same semester and round type.
"""

from __future__ import annotations

import argparse
import json
import os
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import create_engine, text
from sqlalchemy.orm import Session

DEFAULT_DATABASE_URL = "postgresql+psycopg://scheduler:scheduler@localhost:5432/scheduler"
DEFAULT_SEMESTER_CODE = "SU26"
ELIGIBLE_RESULT = "Bảo vệ kỳ 2"
ROUND_TYPE = "DEFENSE_1_2"
ROUND_NAME = "Bảo vệ kỳ 2 (Defense 1.2)"
REQUIRED_HEADERS = {
    "project_code": "mã đề tài",
    "result": "kết quả review 1.1",
}


@dataclass(frozen=True)
class EligibleProject:
    code: str
    source_row: int


def _text(value: Any) -> str:
    return "" if value is None else str(value).replace("\n", " ").strip()


def _header(value: Any) -> str:
    return re.sub(r"\s+", " ", _text(value)).casefold()


def _find_columns(rows: list[tuple[Any, ...]]) -> tuple[int, dict[str, int]]:
    for row_index, row in enumerate(rows, start=1):
        positions = {_header(value): column for column, value in enumerate(row)}
        if all(header in positions for header in REQUIRED_HEADERS.values()):
            return row_index, {
                field: positions[header] for field, header in REQUIRED_HEADERS.items()
            }
    raise ValueError(
        "Could not find a sheet header containing Mã đề tài and Kết quả Review 1.1."
    )


def read_eligible_projects(
    workbook_path: Path,
    requested_sheet: str | None = None,
) -> tuple[str, list[EligibleProject], list[str]]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    candidates = workbook.worksheets
    if requested_sheet:
        try:
            candidates = [workbook[requested_sheet]]
        except KeyError as exc:
            raise ValueError(f"Sheet not found: {requested_sheet}") from exc

    selected_sheet: str | None = None
    header_row: int | None = None
    columns: dict[str, int] | None = None
    rows: list[tuple[Any, ...]] = []
    for worksheet in candidates:
        candidate_rows = list(worksheet.iter_rows(values_only=True))
        try:
            candidate_header_row, candidate_columns = _find_columns(candidate_rows)
        except ValueError:
            continue
        selected_sheet = worksheet.title
        header_row = candidate_header_row
        columns = candidate_columns
        rows = candidate_rows
        break

    if selected_sheet is None or header_row is None or columns is None:
        raise ValueError(
            "No compatible defense eligibility sheet found. "
            f"Available sheets: {', '.join(workbook.sheetnames)}"
        )

    projects: dict[str, EligibleProject] = {}
    for source_row, row in enumerate(rows[header_row:], start=header_row + 1):
        code = _text(row[columns["project_code"]])
        result = _text(row[columns["result"]])
        if not code and not result:
            continue
        if result != ELIGIBLE_RESULT:
            continue
        if not code:
            raise ValueError(f"Eligible row {source_row} has no project code.")
        projects.setdefault(code, EligibleProject(code, source_row))

    if not projects:
        raise ValueError(f"Sheet {selected_sheet!r} contains no eligible projects.")
    return selected_sheet, list(projects.values()), list(workbook.sheetnames)


def load_reference(
    session: Session,
    semester_code: str,
    project_codes: list[str],
) -> dict[str, Any]:
    semester = session.execute(
        text(
            "SELECT id, start_date, end_date, status FROM semesters WHERE code = :code"
        ),
        {"code": semester_code},
    ).mappings().one_or_none()
    if semester is None:
        raise ValueError(f"Semester does not exist: {semester_code}")
    if str(semester["status"]) != "ACTIVE":
        raise ValueError(f"Semester {semester_code} must be ACTIVE to create a round.")

    rows = session.execute(
        text(
            "SELECT p.code AS project_code, g.id AS group_id, g.code AS group_code, "
            "COUNT(gm.student_id) FILTER (WHERE gm.status = 'ACTIVE')::int AS member_count, "
            "COUNT(*) FILTER (WHERE gm.status = 'ACTIVE' AND gm.membership_role = 'LEADER')::int AS leader_count "
            "FROM projects p "
            "LEFT JOIN groups g ON g.project_id = p.id "
            "LEFT JOIN group_memberships gm ON gm.group_id = g.id "
            "WHERE p.semester_id = :semester_id AND p.code = ANY(:project_codes) "
            "GROUP BY p.code, g.id, g.code "
            "ORDER BY p.code"
        ),
        {"semester_id": semester["id"], "project_codes": project_codes},
    ).mappings().all()

    by_project = {str(row["project_code"]): dict(row) for row in rows}
    missing_group_codes = [
        code
        for code in project_codes
        if code not in by_project or by_project[code]["group_id"] is None
    ]
    invalid_group_codes = [
        code
        for code in project_codes
        if code in by_project
        and by_project[code]["group_id"] is not None
        and not (
            int(by_project[code]["member_count"]) <= 5
            and int(by_project[code]["leader_count"]) == 1
        )
    ]
    usable = [
        by_project[code]
        for code in project_codes
        if code in by_project
        and by_project[code]["group_id"] is not None
        and code not in invalid_group_codes
    ]
    return {
        "semester": dict(semester),
        "groups_by_project": by_project,
        "usable_groups": usable,
        "missing_group_codes": missing_group_codes,
        "invalid_group_codes": invalid_group_codes,
    }


def apply_import(
    session: Session,
    reference: dict[str, Any],
    semester_code: str,
    source_name: str,
    eligible_count: int,
    source_sheet: str,
    *,
    attach_round: bool = False,
) -> dict[str, Any]:
    semester = reference["semester"]
    round_id: int | None = None
    round_created = False
    if attach_round:
        round_row = session.execute(
            text(
                "SELECT id, status FROM rounds "
                "WHERE semester_id = :semester_id AND type = CAST(:type AS round_type) "
                "ORDER BY id DESC LIMIT 1 FOR UPDATE"
            ),
            {"semester_id": semester["id"], "type": ROUND_TYPE},
        ).mappings().one_or_none()

        if round_row is not None and str(round_row["status"]) != "DRAFT":
            raise ValueError(
                f"Existing {ROUND_TYPE} round {round_row['id']} is "
                f"{round_row['status']}; refusing to change a non-draft round."
            )

        if round_row is None:
            creator_id = session.execute(
                text(
                    "SELECT a.id FROM accounts a "
                    "JOIN account_roles ar ON ar.account_id = a.id "
                    "WHERE ar.role = 'ADMIN' ORDER BY a.id LIMIT 1"
                )
            ).scalar_one_or_none()
            round_id = session.execute(
                text(
                    "INSERT INTO rounds "
                    "(semester_id, name, description, type, reviewer_count, "
                    "result_owner_mode, group_selection_mode, session_duration_minutes, "
                    "start_date, end_date, created_by) "
                    "VALUES (:semester_id, :name, :description, CAST(:type AS round_type), 5, "
                    "FALSE, FALSE, 30, :start_date, :end_date, :created_by) RETURNING id"
                ),
                {
                    "semester_id": semester["id"],
                    "name": ROUND_NAME,
                    "description": (
                        f"Imported from {source_name}; sheet {source_sheet}; "
                        f"Review 1.1 result = {ELIGIBLE_RESULT}."
                    ),
                    "type": ROUND_TYPE,
                    "start_date": semester["start_date"],
                    "end_date": semester["end_date"],
                    "created_by": creator_id,
                },
            ).scalar_one()
            round_created = True
        else:
            round_id = int(round_row["id"])

    linked = 0
    status_updated = 0
    for group in reference["usable_groups"]:
        if round_id is not None:
            inserted = session.execute(
                text(
                    "INSERT INTO round_groups (round_id, group_id) VALUES (:round_id, :group_id) "
                    "ON CONFLICT DO NOTHING RETURNING group_id"
                ),
                {"round_id": round_id, "group_id": group["group_id"]},
            ).scalar_one_or_none()
            linked += int(inserted is not None)
        updated = session.execute(
            text(
                "UPDATE groups SET status = 'ELIGIBLE_D12' "
                "WHERE id = :group_id AND status = 'PENDING_D11' RETURNING id"
            ),
            {"group_id": group["group_id"]},
        ).scalar_one_or_none()
        status_updated += int(updated is not None)

    session.execute(
        text(
            "INSERT INTO audit_events "
            "(actor_id, action, entity_type, entity_id, reason, after_json) "
            "VALUES (NULL, 'EXCEL_DEFENSE_ELIGIBILITY_IMPORTED', :entity_type, :entity_id, "
            ":reason, CAST(:after_json AS JSONB))"
        ),
        {
            "entity_type": "round" if round_id is not None else "semester",
            "entity_id": str(round_id if round_id is not None else semester["id"]),
            "reason": "Import Defense 1.2 eligible groups from SU26 workbook",
            "after_json": json.dumps(
                {
                    "semester": semester_code,
                    "source": source_name,
                    "sheet": source_sheet,
                    "eligible_projects": eligible_count,
                    "usable_groups": len(reference["usable_groups"]),
                    "round_id": round_id,
                    "linked_groups": linked,
                    "status_updated": status_updated,
                }
            ),
        },
    )
    return {
        "round_id": round_id,
        "round_created": round_created,
        "groups_linked": linked,
        "new_links": linked,
        "group_status_updated": status_updated,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--sheet")
    parser.add_argument("--semester", default=DEFAULT_SEMESTER_CODE)
    parser.add_argument(
        "--database-url",
        default=os.getenv("DATABASE_URL", DEFAULT_DATABASE_URL),
    )
    parser.add_argument("--apply", action="store_true")
    parser.add_argument(
        "--attach-round",
        action="store_true",
        help="Also create/reuse the Defense 1.2 round and attach eligible groups.",
    )
    args = parser.parse_args()

    workbook_path = args.workbook.resolve()
    if not workbook_path.is_file():
        raise SystemExit(f"Workbook not found: {workbook_path}")

    try:
        sheet_name, projects, sheet_names = read_eligible_projects(
            workbook_path, args.sheet
        )
        engine = create_engine(args.database_url, pool_pre_ping=True)
        with Session(engine) as session:
            reference = load_reference(
                session, args.semester, [project.code for project in projects]
            )
            result: dict[str, Any] = {
                "mode": "apply" if args.apply else "dry-run",
                "workbook": workbook_path.name,
                "sheets": sheet_names,
                "source_sheet": sheet_name,
                "eligible_projects": len(projects),
                "groups_found": len(projects)
                - len(reference["missing_group_codes"]),
                "usable_groups": len(reference["usable_groups"]),
                "missing_group_count": len(reference["missing_group_codes"]),
                "invalid_group_count": len(reference["invalid_group_codes"]),
                "missing_group_codes": reference["missing_group_codes"],
                "invalid_group_codes": reference["invalid_group_codes"],
            }
            if args.apply:
                session.rollback()
                with session.begin():
                    result.update(
                        apply_import(
                            session,
                            reference,
                            args.semester,
                            workbook_path.name,
                            len(projects),
                            sheet_name,
                            attach_round=args.attach_round,
                        )
                    )
            print(json.dumps(result, ensure_ascii=False, indent=2, default=str))
    except (OSError, ValueError) as exc:
        raise SystemExit(str(exc)) from exc


if __name__ == "__main__":
    main()
