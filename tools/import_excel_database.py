"""Replace local scheduler data with normalized data parsed from an Excel workbook.

The workbook binary is never stored.  Every non-empty sheet row is stored as JSONB,
and the fields that fit the scheduler domain are also projected into canonical tables.
Run from the repository root with the local Docker PostgreSQL service running.
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import math
import os
import random
import re
import subprocess
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

DEMO_PASSWORD_HASH = (
    "$argon2id$v=19$m=65536,t=3,p=4$V6pAiT9HmR/z0IItWjtNuA$"
    "skwmr/puZ9G53gYN9UBRlHgGHXIVk2nNuIh9OunS00Y"
)
SEMESTER_CODE = "SE-2026-2027"
MAJOR_CODE = "SE"


def normalize(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return None
        if value.is_integer():
            return int(value)
    return value


def key(value: Any) -> str | None:
    value = normalize(value)
    if value is None or value == "":
        return None
    return str(value)


def integer(value: Any) -> int | None:
    value = normalize(value)
    if value is None or value == "":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def date_value(value: Any) -> str | None:
    value = normalize(value)
    if value is None:
        return None
    return str(value)[:10]


def clean_code(value: Any) -> str | None:
    value = key(value)
    if value is None or value.startswith("#"):
        return None
    return value


def sql(value: Any) -> str:
    value = normalize(value)
    if value is None:
        return "NULL"
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (int, float)):
        return str(value)
    return "'" + str(value).replace("'", "''") + "'"


def json_sql(value: Any) -> str:
    payload = json.dumps(value, ensure_ascii=False, separators=(",", ":"), default=str)
    return sql(payload) + "::jsonb"


def rows_for_sheet(values_ws: Any, formulas_ws: Any) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    value_rows = values_ws.iter_rows(values_only=True)
    formula_rows = formulas_ws.iter_rows(values_only=True)
    for row_number, (value_row, formula_row) in enumerate(zip(value_rows, formula_rows), start=1):
        values = [normalize(value) for value in value_row]
        formulas = {
            str(column): formula
            for column, formula in enumerate(formula_row)
            if isinstance(formula, str) and formula.startswith("=")
        }
        if any(value is not None and value != "" for value in values) or formulas:
            rows.append({"row_number": row_number, "values": values, "formulas": formulas})
    return rows


def slug(value: str) -> str:
    result = value.encode("ascii", "ignore").decode("ascii").lower()
    result = re.sub(r"[^a-z0-9]+", "_", result).strip("_")
    return result or "lecturer"


def add_sql(lines: list[str], statement: str) -> None:
    lines.append(statement.rstrip(";") + ";")


def psycopg_url(database_url: str) -> str:
    return database_url.replace("postgresql+psycopg://", "postgresql://", 1)


def database_has_data(database_url: str) -> bool:
    import psycopg

    with psycopg.connect(psycopg_url(database_url)) as connection:
        return bool(
            connection.execute(
                "SELECT EXISTS ("
                "SELECT 1 FROM excel_import_batches UNION ALL "
                "SELECT 1 FROM accounts UNION ALL "
                "SELECT 1 FROM projects UNION ALL "
                "SELECT 1 FROM groups UNION ALL "
                "SELECT 1 FROM sessions"
                ")"
            ).fetchone()[0]
        )


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--database-url", default=os.getenv("DATABASE_URL"))
    parser.add_argument("--skip-if-imported", action="store_true")
    parser.add_argument("--sample-fraction", type=float, default=1.0, help="Keep only this fraction of projects (and their linked review/defense rows).")
    parser.add_argument("--sample-seed", type=int, default=42)
    parser.add_argument("--output", type=Path, help="Also write the generated SQL to this file.")
    parser.add_argument("--apply", action="store_true", help="Actually run the generated SQL against the database (destructive: truncates all tables first). Without this flag, the script only writes --output.")
    args = parser.parse_args()
    if not 0 < args.sample_fraction <= 1.0:
        raise SystemExit("--sample-fraction must be in (0, 1]")
    workbook_path = args.workbook.resolve()
    if not workbook_path.is_file():
        raise SystemExit(f"Workbook not found: {workbook_path}")
    if args.skip_if_imported:
        if not args.database_url:
            raise SystemExit("--skip-if-imported requires DATABASE_URL or --database-url")
        if database_has_data(args.database_url):
            print(
                json.dumps(
                    {
                        "source": str(workbook_path),
                        "status": "skipped",
                        "reason": "database_already_populated",
                    }
                )
            )
            return

    values_book = load_workbook(workbook_path, data_only=True, read_only=True)
    formulas_book = load_workbook(workbook_path, data_only=False, read_only=True)
    sheet_rows = {
        name: rows_for_sheet(values_book[name], formulas_book[name]) for name in values_book.sheetnames
    }

    project_rows = [
        row
        for row in sheet_rows["Project"]
        if row["row_number"] >= 4 and (row["values"][2] or row["values"][3])
    ]
    if args.sample_fraction < 1.0:
        sample_size = max(1, round(len(project_rows) * args.sample_fraction))
        sampled = random.Random(args.sample_seed).sample(project_rows, k=sample_size)
        project_rows = sorted(sampled, key=lambda row: row["row_number"])

    project_by_review_code: dict[tuple[str, str], dict[str, Any]] = {}
    project_by_council_code: dict[tuple[str, str], list[dict[str, Any]]] = {
        ("Defense1", ""): [],
        ("Defense2", ""): [],
    }
    for project in project_rows:
        values = project["values"]
        for review_type, column in (("Review1", 9), ("Review2", 19), ("Review3", 29)):
            schedule_code = clean_code(values[column])
            if schedule_code:
                project_by_review_code[(review_type, schedule_code)] = project
        for defense_type, column in (("Defense1", 41), ("Defense2", 51)):
            council_code = clean_code(values[column])
            if council_code:
                project_by_council_code.setdefault((defense_type, council_code), []).append(project)

    review_rows: dict[str, list[dict[str, Any]]] = {}
    for sheet_name in ("Review1", "Review2", "Review3"):
        review_rows[sheet_name] = [
            row
            for row in sheet_rows[sheet_name]
            if row["row_number"] >= 5 and clean_code(row["values"][0])
        ]

    defense_rows: dict[str, list[dict[str, Any]]] = {}
    for sheet_name in ("Defense1", "Defense2"):
        defense_rows[sheet_name] = [
            row
            for row in sheet_rows[sheet_name]
            if row["row_number"] >= 3 and clean_code(row["values"][3])
        ]

    summary_rows = [
        row
        for row in sheet_rows["Summary"]
        if row["row_number"] >= 2 and clean_code(row["values"][1])
    ]

    lecturer_codes: set[str] = set()
    for project in project_rows:
        lecturer_codes.update(
            code for code in (clean_code(project["values"][7]), clean_code(project["values"][8])) if code
        )
    for rows in review_rows.values():
        for row in rows:
            lecturer_codes.update(code for code in (clean_code(row["values"][9]), clean_code(row["values"][10])) if code)
    for rows in defense_rows.values():
        for row in rows:
            lecturer_codes.update(code for code in (clean_code(value) for value in row["values"][4:9]) if code)
    lecturer_codes.update(clean_code(row["values"][1]) for row in summary_rows if clean_code(row["values"][1]))

    rooms = sorted(
        {
            clean_code(row["values"][8])
            for rows in review_rows.values()
            for row in rows
            if clean_code(row["values"][8])
        }
    )

    lines: list[str] = ["BEGIN;", "SET TIME ZONE 'Asia/Ho_Chi_Minh';"]
    lines.append(
        "DO $$ DECLARE statement TEXT; BEGIN "
        "SELECT 'TRUNCATE TABLE ' || string_agg(format('%I.%I', table_schema, table_name), ', ') "
        "|| ' RESTART IDENTITY CASCADE' INTO statement FROM information_schema.tables "
        "WHERE table_schema = 'public' AND table_type = 'BASE TABLE' "
        "AND table_name NOT IN ('alembic_version', 'schema_meta'); "
        "EXECUTE statement; END $$;"
    )

    add_sql(
        lines,
        "INSERT INTO semesters (code, name, note, start_date, end_date, academic_year, status) VALUES "
        f"({sql(SEMESTER_CODE)}, {sql('Excel import: SE_CapstoneProject_SP26_ReviewDefense_New')}, "
        f"{sql('Excel import seed')}, DATE '2026-05-11', DATE '2026-08-23', {sql('2026-2027')}, 'ACTIVE')",
    )
    add_sql(
        lines,
        "INSERT INTO majors (code, name) VALUES "
        f"({sql(MAJOR_CODE)}, {sql('Software Engineering')})",
    )
    for email, display_name, role in (
        ("admin1@gmail.com", "Scheduler Admin 1", "ADMIN"),
        ("admin2@gmail.com", "Scheduler Admin 2", "ADMIN"),
        ("manager1@gmail.com", "Scheduler Manager 1", "MANAGER"),
        ("manager2@gmail.com", "Scheduler Manager 2", "MANAGER"),
    ):
        add_sql(
            lines,
            "INSERT INTO accounts (email, display_name, password_hash) VALUES "
            f"({sql(email)}, {sql(display_name)}, {sql(DEMO_PASSWORD_HASH)})",
        )
        add_sql(
            lines,
            "INSERT INTO account_roles (account_id, role) SELECT id, "
            f"{sql(role)}::system_role FROM accounts WHERE email = {sql(email)}",
        )

    for lecturer_number, lecturer_code in enumerate(sorted(lecturer_codes), start=1):
        email = (
            f"lecturer{lecturer_number}@gmail.com"
            if lecturer_number <= 2
            else f"{slug(lecturer_code)}@excel.local"
        )
        add_sql(
            lines,
            "INSERT INTO accounts (email, display_name, password_hash) VALUES "
            f"({sql(email)}, {sql(lecturer_code)}, {sql(DEMO_PASSWORD_HASH)})",
        )
        add_sql(
            lines,
            "INSERT INTO account_roles (account_id, role) SELECT id, 'LECTURER'::system_role "
            f"FROM accounts WHERE email = {sql(email)}",
        )
        add_sql(
            lines,
            "INSERT INTO lecturers (account_id, lecturer_code) SELECT id, "
            f"{sql(lecturer_code)} FROM accounts WHERE email = {sql(email)}",
        )

    for student_number in range(1, 3):
        email = f"student{student_number}@gmail.com"
        student_code = f"SV{student_number:03d}"
        add_sql(
            lines,
            "INSERT INTO accounts (email, display_name, password_hash) VALUES "
            f"({sql(email)}, {sql(f'Sinh viên {student_number:03d}')}, {sql(DEMO_PASSWORD_HASH)})",
        )
        add_sql(
            lines,
            "INSERT INTO account_roles (account_id, role) SELECT id, 'STUDENT'::system_role "
            f"FROM accounts WHERE email = {sql(email)}",
        )
        add_sql(
            lines,
            "INSERT INTO students (account_id, student_code) SELECT id, "
            f"{sql(student_code)} FROM accounts WHERE email = {sql(email)}",
        )

    for room in rooms:
        add_sql(
            lines,
            "INSERT INTO rooms (code, name, capacity) VALUES "
            f"({sql(room)}, {sql(room)}, 999)",
        )

    round_specs = {
        "Review1": ("REVIEW_1", 2),
        "Review2": ("REVIEW_2", 2),
        "Review3": ("REVIEW_3", 2),
        "Defense1": ("DEFENSE_1", 5),
        "Defense2": ("DEFENSE_2", 5),
    }
    admin_id = "(SELECT id FROM accounts WHERE email = 'admin1@gmail.com')"
    for round_type, reviewer_count in round_specs.values():
        add_sql(
            lines,
            "INSERT INTO rounds (semester_id, type, status, session_duration_minutes, reviewer_count, "
            "result_owner_mode, created_by) SELECT id, "
            f"{sql(round_type)}::round_type, 'PUBLISHED'::round_status, 30, {reviewer_count}, FALSE, {admin_id} "
            f"FROM semesters WHERE code = {sql(SEMESTER_CODE)}",
        )

    batch_lookup = "(SELECT id FROM excel_import_batches ORDER BY id DESC LIMIT 1)"
    add_sql(
        lines,
        "INSERT INTO excel_import_batches (source_file_name, source_path, notes) VALUES "
        f"({sql(workbook_path.name)}, {sql(str(workbook_path))}, "
        f"{sql('Parsed workbook; binary file not stored')})",
    )

    for sheet_name, rows in sheet_rows.items():
        for row in rows:
            add_sql(
                lines,
                "INSERT INTO excel_sheet_rows "
                "(batch_id, sheet_name, row_number, values_jsonb, formulas_jsonb, non_empty) VALUES "
                f"({batch_lookup}, {sql(sheet_name)}, {row['row_number']}, "
                f"{json_sql(row['values'])}, {json_sql(row['formulas'])}, TRUE)",
            )

    semester_lookup = f"(SELECT id FROM semesters WHERE code = {sql(SEMESTER_CODE)})"
    major_lookup = f"(SELECT id FROM majors WHERE code = {sql(MAJOR_CODE)})"
    for project in project_rows:
        values = project["values"]
        project_code = clean_code(values[2])
        group_code = clean_code(values[3])
        if not project_code or not group_code:
            continue
        title_en = str(values[4] or values[5] or project_code)[:255]
        title_vi = values[5]
        add_sql(
            lines,
            "INSERT INTO projects (semester_id, major_id, code, title) VALUES "
            f"({semester_lookup}, {major_lookup}, {sql(project_code)}, {sql(title_en)})",
        )
        add_sql(
            lines,
            "INSERT INTO groups (project_id, code) SELECT id, "
            f"{sql(group_code)} FROM projects WHERE semester_id = {semester_lookup} "
            f"AND code = {sql(project_code)}",
        )
        add_sql(
            lines,
            "INSERT INTO excel_projects "
            "(batch_id, excel_row, project_code, group_code, title_en, title_vi, supervisor_display_name, "
            "supervisor_1_code, supervisor_2_code, canonical_project_id, canonical_group_id, raw_values) "
            "SELECT "
            f"{batch_lookup}, {project['row_number']}, {sql(project_code)}, {sql(group_code)}, "
            f"{sql(values[4])}, {sql(title_vi)}, {sql(values[6])}, {sql(clean_code(values[7]))}, "
            f"{sql(clean_code(values[8]))}, p.id, g.id, {json_sql(values)} "
            "FROM projects p JOIN groups g ON g.project_id = p.id "
            f"WHERE p.semester_id = {semester_lookup} AND p.code = {sql(project_code)}",
        )
        for lecturer_code, supervisor_type in (
            (clean_code(values[7]), "MAIN"),
            (clean_code(values[8]), "CO"),
        ):
            if lecturer_code:
                add_sql(
                    lines,
                    "INSERT INTO project_supervisors (project_id, lecturer_id, supervisor_type) "
                    f"SELECT p.id, l.id, {sql(supervisor_type)}::supervisor_type FROM projects p "
                    f"JOIN lecturers l ON l.lecturer_code = {sql(lecturer_code)} "
                    f"WHERE p.semester_id = {semester_lookup} AND p.code = {sql(project_code)} "
                    "ON CONFLICT (project_id, lecturer_id) DO UPDATE SET supervisor_type = EXCLUDED.supervisor_type",
                )

    first_project = next(
        (
            project
            for project in project_rows
            if clean_code(project["values"][2]) and clean_code(project["values"][3])
        ),
        None,
    )
    if first_project:
        first_project_code = clean_code(first_project["values"][2])
        first_group_code = clean_code(first_project["values"][3])
        for student_number, membership_role in ((1, "LEADER"), (2, "MEMBER")):
            add_sql(
                lines,
                "INSERT INTO group_memberships (group_id, student_id, membership_role) "
                "SELECT g.id, s.id, "
                f"{sql(membership_role)}::membership_role FROM groups g "
                "JOIN projects p ON p.id = g.project_id "
                f"JOIN students s ON s.student_code = {sql(f'SV{student_number:03d}')} "
                f"WHERE p.semester_id = {semester_lookup} AND p.code = {sql(first_project_code)} "
                f"AND g.code = {sql(first_group_code)}",
            )

    # Review sheets contain complete date/slot/room/reviewer data, so project them into sessions.
    for sheet_name, rows in review_rows.items():
        round_type = round_specs[sheet_name][0]
        for row in rows:
            values = row["values"]
            schedule_code = clean_code(values[0])
            project = project_by_review_code.get((sheet_name, schedule_code or ""))
            if not project:
                continue
            project_code = clean_code(project["values"][2])
            group_code = clean_code(project["values"][3])
            schedule_date = date_value(values[6])
            slot_number = integer(values[3]) or 1
            start_time = f"{schedule_date}T{8 + slot_number:02d}:00:00+07:00" if schedule_date else None
            end_time = f"{schedule_date}T{8 + slot_number:02d}:30:00+07:00" if schedule_date else None
            round_lookup = f"(SELECT id FROM rounds WHERE semester_id = {semester_lookup} AND type = {sql(round_type)}::round_type)"
            if schedule_date and start_time and end_time:
                add_sql(
                    lines,
                    "INSERT INTO round_days (round_id, day_date) VALUES "
                    f"({round_lookup}, {sql(schedule_date)}) ON CONFLICT (round_id, day_date) DO NOTHING",
                )
                add_sql(
                    lines,
                    "INSERT INTO timeslots (round_day_id, start_at, end_at) SELECT rd.id, "
                    f"{sql(start_time)}::timestamptz, {sql(end_time)}::timestamptz FROM round_days rd "
                    f"WHERE rd.round_id = {round_lookup} AND rd.day_date = {sql(schedule_date)} "
                    "ON CONFLICT (round_day_id, start_at, end_at) DO NOTHING",
                )
            if values[8]:
                add_sql(
                    lines,
                    "INSERT INTO round_room_types (round_id, room_type) SELECT "
                    f"{round_lookup}, room_type FROM rooms WHERE code = {sql(clean_code(values[8]))} "
                    "ON CONFLICT DO NOTHING",
                )
            add_sql(
                lines,
                "INSERT INTO round_groups (round_id, group_id) SELECT "
                f"{round_lookup}, g.id FROM groups g JOIN projects p ON p.id = g.project_id "
                f"WHERE p.semester_id = {semester_lookup} AND p.code = {sql(project_code)} "
                "ON CONFLICT DO NOTHING",
            )
            if schedule_date and start_time and end_time:
                add_sql(
                    lines,
                    "INSERT INTO schedule_versions (round_id, version_no, status, input_snapshot, created_by) "
                    "SELECT "
                    f"{round_lookup}, 1, 'PUBLISHED'::schedule_version_status, "
                    f"{json_sql({'source': 'EXCEL_IMPORT', 'sheet': sheet_name})}, {admin_id} "
                    "WHERE NOT EXISTS (SELECT 1 FROM schedule_versions WHERE round_id = "
                    f"{round_lookup} AND version_no = 1)",
                )
                # A Session must reference a sealed Council (migration 0025_immutable_councils);
                # build and seal one per review row before the Session row can be inserted.
                council_reason = sql(f"excel:{sheet_name}:{row['row_number']}")
                council_lookup = f"(SELECT id FROM councils WHERE round_id = {round_lookup} AND reason = {council_reason})"
                add_sql(
                    lines,
                    "INSERT INTO councils (round_id, created_by, reason) SELECT "
                    f"{round_lookup}, {admin_id}, {council_reason} "
                    f"WHERE NOT EXISTS (SELECT 1 FROM councils WHERE round_id = {round_lookup} AND reason = {council_reason})",
                )
                for lecturer_code in (clean_code(values[9]), clean_code(values[10])):
                    if lecturer_code:
                        add_sql(
                            lines,
                            "INSERT INTO council_members (council_id, lecturer_id, assignment, snapshot_name) "
                            f"SELECT {council_lookup}, l.id, 'REVIEWER'::assignment_role, l.lecturer_code "
                            f"FROM lecturers l WHERE l.lecturer_code = {sql(lecturer_code)} "
                            f"AND NOT EXISTS (SELECT 1 FROM council_members WHERE council_id = {council_lookup} AND lecturer_id = l.id)",
                        )
                add_sql(
                    lines,
                    f"UPDATE councils SET sealed_at = now() WHERE id = {council_lookup} AND sealed_at IS NULL",
                )
                add_sql(
                    lines,
                    "INSERT INTO sessions "
                    "(schedule_version_id, group_id, timeslot_id, room_id, council_id, start_at, end_at, status) "
                    "SELECT sv.id, g.id, ts.id, rm.id, "
                    f"{council_lookup}, "
                    f"{sql(start_time)}::timestamptz, {sql(end_time)}::timestamptz, 'SCHEDULED'::session_status "
                    "FROM schedule_versions sv JOIN groups g ON g.code = "
                    f"{sql(group_code)} JOIN projects p ON p.id = g.project_id AND p.semester_id = {semester_lookup} "
                    "JOIN round_days rd ON rd.round_id = sv.round_id AND rd.day_date = "
                    f"{sql(schedule_date)} JOIN timeslots ts ON ts.round_day_id = rd.id AND ts.start_at = "
                    f"{sql(start_time)}::timestamptz AND ts.end_at = {sql(end_time)}::timestamptz "
                    f"JOIN rooms rm ON rm.code = {sql(clean_code(values[8]))} "
                    f"WHERE sv.round_id = {round_lookup} AND sv.version_no = 1 "
                    "ON CONFLICT (schedule_version_id, group_id) DO NOTHING",
                )
            add_sql(
                lines,
                "INSERT INTO excel_review_schedule_rows "
                "(batch_id, review_type, excel_row, schedule_code, week_code, day_code, slot_number, wds_code, "
                "group_number, schedule_date, date_of_week, room_name, reviewer_1_code, reviewer_2_code, count_value, "
                "canonical_round_id, canonical_session_id, raw_values) VALUES ("
                f"{batch_lookup}, {sql(sheet_name)}, {row['row_number']}, {sql(schedule_code)}, "
                f"{sql(integer(values[1]))}, {sql(integer(values[2]))}, {sql(integer(values[3]))}, "
                f"{sql(integer(values[4]))}, {sql(integer(values[5]))}, {sql(schedule_date)}, {sql(values[7])}, "
                f"{sql(clean_code(values[8]))}, {sql(clean_code(values[9]))}, {sql(clean_code(values[10]))}, "
                f"{sql(integer(values[11]))}, {round_lookup}, "
                f"(SELECT s.id FROM sessions s JOIN schedule_versions sv ON sv.id = s.schedule_version_id "
                f"JOIN groups g ON g.id = s.group_id WHERE sv.round_id = {round_lookup} AND g.code = {sql(group_code)} LIMIT 1), "
                f"{json_sql(values)})",
            )

    for sheet_name, rows in defense_rows.items():
        round_type = round_specs[sheet_name][0]
        for row in rows:
            values = row["values"]
            council_code = clean_code(values[3])
            council_date = date_value(values[1])
            members = [clean_code(value) for value in values[4:9]]
            members = [member for member in members if member]
            round_lookup = f"(SELECT id FROM rounds WHERE semester_id = {semester_lookup} AND type = {sql(round_type)}::round_type)"
            if council_date:
                add_sql(
                    lines,
                    "INSERT INTO round_days (round_id, day_date) VALUES "
                    f"({round_lookup}, {sql(council_date)}) ON CONFLICT (round_id, day_date) DO NOTHING",
                )
            add_sql(
                lines,
                "INSERT INTO excel_defense_councils "
                "(batch_id, defense_type, excel_row, council_code, council_date, day_code, chair_code, secretary_code, "
                "member_1_code, member_2_code, member_3_code, member_count, group_count, member_list, canonical_round_id, raw_values) VALUES ("
                f"{batch_lookup}, {sql(sheet_name)}, {row['row_number']}, {sql(council_code)}, {sql(council_date)}, "
                f"{sql(integer(values[2]))}, {sql(clean_code(values[4]))}, {sql(clean_code(values[5]))}, "
                f"{sql(clean_code(values[6]))}, {sql(clean_code(values[7]))}, {sql(clean_code(values[8]))}, "
                f"{len(members)}, {sql(integer(values[10]))}, {sql(values[9])}, {round_lookup}, {json_sql(values)})",
            )
            for project in project_by_council_code.get((sheet_name, council_code or ""), []):
                project_code = clean_code(project["values"][2])
                group_code = clean_code(project["values"][3])
                if not project_code or not group_code:
                    continue
                add_sql(
                    lines,
                    "INSERT INTO round_groups (round_id, group_id) SELECT "
                    f"{round_lookup}, g.id FROM groups g JOIN projects p ON p.id = g.project_id "
                    f"WHERE p.semester_id = {semester_lookup} AND p.code = {sql(project_code)} "
                    "ON CONFLICT DO NOTHING",
                )
                add_sql(
                    lines,
                    "INSERT INTO excel_council_groups (council_id, project_code, group_code, project_id, group_id) "
                    "SELECT c.id, "
                    f"{sql(project_code)}, {sql(group_code)}, p.id, g.id FROM excel_defense_councils c "
                    f"JOIN projects p ON p.code = {sql(project_code)} AND p.semester_id = {semester_lookup} "
                    f"JOIN groups g ON g.project_id = p.id AND g.code = {sql(group_code)} "
                    f"WHERE c.batch_id = {batch_lookup} AND c.defense_type = {sql(sheet_name)} "
                    f"AND c.council_code = {sql(council_code)} ON CONFLICT DO NOTHING",
                )

    for row in summary_rows:
        values = row["values"]
        add_sql(
            lines,
            "INSERT INTO excel_summary_workloads "
            "(batch_id, excel_row, lecturer_code, department, review_1_count, review_2_count, review_3_count, "
            "defense_1_count, defense_2_count, raw_values) VALUES ("
            f"{batch_lookup}, {row['row_number']}, {sql(clean_code(values[1]))}, {sql(values[2])}, "
            f"{sql(integer(values[4]))}, {sql(integer(values[5]))}, {sql(integer(values[6]))}, "
            f"{sql(integer(values[8]))}, {sql(integer(values[9]))}, {json_sql(values)})",
        )

    lines.append("COMMIT;")
    sql_payload = "\n".join(lines) + "\n"
    if args.output:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(sql_payload, encoding="utf-8")
    if args.apply:
        if args.database_url:
            import psycopg

            with psycopg.connect(psycopg_url(args.database_url), autocommit=True) as connection:
                connection.execute(sql_payload)
        else:
            command = [
                "docker",
                "compose",
                "exec",
                "-T",
                "postgres",
                "psql",
                "-U",
                "scheduler",
                "-d",
                "scheduler",
                "-v",
                "ON_ERROR_STOP=1",
                "-X",
                "-q",
            ]
            subprocess.run(
                command,
                input=sql_payload,
                text=True,
                encoding="utf-8",
                check=True,
                cwd=Path.cwd(),
            )
    print(
        json.dumps(
            {
                "source": str(workbook_path),
                "sample_fraction": args.sample_fraction,
                "output": str(args.output) if args.output else None,
                "sheets": list(sheet_rows),
                "projects": len(project_rows),
                "lecturers": len(lecturer_codes),
                "rooms": len(rooms),
                "review_rows": {name: len(rows) for name, rows in review_rows.items()},
                "defense_rows": {name: len(rows) for name, rows in defense_rows.items()},
                "summary_rows": len(summary_rows),
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
