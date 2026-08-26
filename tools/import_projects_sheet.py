"""Import project titles and supervisors from the one-sheet SU26 workbook.

The importer is deliberately separate from ``import_excel_database.py`` because
that script targets the older multi-sheet FAP workbook format.  This command is
safe to run repeatedly: it upserts projects for one semester and reconciles
their supervisor assignments, without truncating any table or creating accounts.

Default mode is validation-only.  Pass ``--apply`` to write to PostgreSQL.
"""

from __future__ import annotations

import argparse
import json
import os
import re
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import create_engine, text
from sqlalchemy.orm import Session

DEFAULT_DATABASE_URL = "postgresql+psycopg://scheduler:scheduler@localhost:5432/scheduler"
DEFAULT_SEMESTER_CODE = "SU26"
DEFAULT_MAJOR_CODE = "SE"
REQUIRED_HEADERS = {
    "project_code": "mã đề tài",
    "title_vi": "tên đề tài tiếng việt",
    "title_en": "tên đề tài tiếng anh/ tiếng nhật",
    "supervisors": "gvhd",
}
OPTIONAL_HEADERS = {
    "topic_type": ("loại đề tài", "topic type", "topictype"),
}


@dataclass(frozen=True)
class ProjectRow:
    code: str
    title_vi: str
    title_en: str
    supervisors: tuple[str, ...]
    topic_type: str
    source_row: int


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()


def _header(value: Any) -> str:
    return re.sub(r"\s+", " ", _text(value)).casefold()


def _supervisor_names(value: Any) -> tuple[str, ...]:
    names = tuple(part.strip() for part in _text(value).split(" và ") if part.strip())
    return names


def _find_header_row(rows: list[tuple[Any, ...]]) -> tuple[int, dict[str, int]]:
    for row_index, row in enumerate(rows, start=1):
        positions = {_header(value): column for column, value in enumerate(row)}
        if all(header in positions for header in REQUIRED_HEADERS.values()):
            columns = {
                field: positions[header] for field, header in REQUIRED_HEADERS.items()
            }
            for field, headers in OPTIONAL_HEADERS.items():
                position = next((positions[header] for header in headers if header in positions), None)
                if position is not None:
                    columns[field] = position
            return row_index, columns
    raise ValueError(
        "Could not find a sheet header containing Mã đề tài, "
        "Tên đề tài Tiếng Việt, Tên đề tài Tiếng Anh/ Tiếng Nhật and GVHD."
    )


def read_projects(workbook_path: Path, requested_sheet: str | None = None) -> tuple[str, list[ProjectRow]]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    candidates = workbook.worksheets
    if requested_sheet:
        try:
            candidates = [workbook[requested_sheet]]
        except KeyError as exc:
            raise ValueError(f"Sheet not found: {requested_sheet}") from exc

    selected_sheet = None
    header_row = None
    columns = None
    rows: list[tuple[Any, ...]] = []
    for worksheet in candidates:
        candidate_rows = list(worksheet.iter_rows(values_only=True))
        try:
            candidate_header_row, candidate_columns = _find_header_row(candidate_rows)
        except ValueError:
            continue
        selected_sheet = worksheet.title
        header_row = candidate_header_row
        columns = candidate_columns
        rows = candidate_rows
        break

    if selected_sheet is None or header_row is None or columns is None:
        names = ", ".join(workbook.sheetnames)
        raise ValueError(f"No compatible project sheet found. Available sheets: {names}")

    projects: dict[str, ProjectRow] = {}
    for source_row, row in enumerate(rows[header_row:], start=header_row + 1):
        code = _text(row[columns["project_code"]])
        title_vi = _text(row[columns["title_vi"]])
        title_en = _text(row[columns["title_en"]])
        supervisors = _supervisor_names(row[columns["supervisors"]])
        topic_type = _text(row[columns["topic_type"]]).upper() if "topic_type" in columns else "REGULAR"
        if not any((code, title_vi, title_en, supervisors)):
            continue
        if not code or not title_vi or not supervisors:
            raise ValueError(
                f"Incomplete project row {source_row}: "
                f"code={code!r}, title_vi={title_vi!r}, supervisors={supervisors!r}"
            )
        if len(code) > 64:
            raise ValueError(f"Project code at row {source_row} exceeds 64 characters: {code}")
        if len(title_vi) > 255:
            raise ValueError(f"Vietnamese title at row {source_row} exceeds 255 characters: {code}")
        if len(title_en) > 255:
            raise ValueError(f"English title at row {source_row} exceeds 255 characters: {code}")
        if len(supervisors) > 2:
            raise ValueError(f"More than two supervisors at row {source_row}: {code}")
        if topic_type not in {"APPLICATION", "RESEARCH", "INTEGRATED", "REGULAR"}:
            raise ValueError(f"Invalid topic type at row {source_row}: {code}")

        project = ProjectRow(code, title_vi, title_en, supervisors, topic_type, source_row)
        previous = projects.get(code)
        if previous is not None and (
            previous.title_vi != project.title_vi
            or previous.title_en != project.title_en
            or previous.supervisors != project.supervisors
            or previous.topic_type != project.topic_type
        ):
            raise ValueError(
                f"Project {code} has inconsistent repeated rows "
                f"({previous.source_row} and {source_row})."
            )
        projects.setdefault(code, project)

    if not projects:
        raise ValueError(f"Sheet {selected_sheet!r} contains no project rows.")
    return selected_sheet, list(projects.values())


def _load_db_reference(session: Session, semester_code: str, major_code: str) -> dict[str, Any]:
    semester = session.execute(
        text("SELECT id FROM semesters WHERE code = :code"), {"code": semester_code}
    ).scalar_one_or_none()
    if semester is None:
        raise ValueError(f"Semester does not exist: {semester_code}")

    major = session.execute(
        text("SELECT id FROM majors WHERE code = :code"), {"code": major_code}
    ).scalar_one_or_none()
    if major is None:
        raise ValueError(f"Major does not exist: {major_code}")

    lecturer_rows = session.execute(
        text(
            "SELECT l.id, l.lecturer_code, a.display_name "
            "FROM lecturers l JOIN accounts a ON a.id = l.account_id "
            "WHERE a.status = 'ACTIVE'"
        )
    ).mappings().all()
    lecturers_by_name: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for lecturer in lecturer_rows:
        lecturers_by_name[_text(lecturer["display_name"])].append(dict(lecturer))
    return {"semester_id": int(semester), "major_id": int(major), "lecturers_by_name": lecturers_by_name}


def validate_projects(
    session: Session,
    projects: list[ProjectRow],
    semester_code: str,
    major_code: str,
) -> dict[str, Any]:
    reference = _load_db_reference(session, semester_code, major_code)
    missing: set[str] = set()
    ambiguous: dict[str, int] = {}
    assignments: dict[str, list[dict[str, Any]]] = {}
    for project in projects:
        project_assignments: list[dict[str, Any]] = []
        for supervisor_name in project.supervisors:
            matches = reference["lecturers_by_name"].get(supervisor_name, [])
            if not matches:
                missing.add(supervisor_name)
            elif len(matches) != 1:
                ambiguous[supervisor_name] = len(matches)
            else:
                project_assignments.append(matches[0])
        assignments[project.code] = project_assignments

    if missing or ambiguous:
        details = []
        if missing:
            details.append(f"missing lecturer names: {sorted(missing)}")
        if ambiguous:
            details.append(f"ambiguous lecturer names: {ambiguous}")
        raise ValueError("GVHD validation failed; " + "; ".join(details))

    existing_codes = set(
        session.execute(
            text(
                "SELECT p.code FROM projects p "
                "JOIN semesters s ON s.id = p.semester_id "
                "WHERE s.code = :semester_code"
            ),
            {"semester_code": semester_code},
        ).scalars()
    )
    return {
        "semester_id": reference["semester_id"],
        "major_id": reference["major_id"],
        "assignments": assignments,
        "existing_projects": len(existing_codes.intersection({project.code for project in projects})),
        "missing_projects": len({project.code for project in projects} - existing_codes),
    }


def apply_projects(
    session: Session,
    projects: list[ProjectRow],
    validation: dict[str, Any],
    semester_code: str,
    source_name: str,
) -> dict[str, int | str]:
    created = 0
    updated = 0
    supervisor_assignments = 0
    for project in projects:
        existed = session.execute(
            text(
                "SELECT 1 FROM projects p JOIN semesters s ON s.id = p.semester_id "
                "WHERE s.code = :semester_code AND p.code = :code"
            ),
            {"semester_code": semester_code, "code": project.code},
        ).scalar_one_or_none()
        project_id = session.execute(
            text(
                "INSERT INTO projects (semester_id, major_id, code, title, title_vi, title_en, topic_type) "
                "VALUES (:semester_id, :major_id, :code, :title, :title_vi, :title_en, CAST(:topic_type AS topic_type)) "
                "ON CONFLICT (semester_id, code) DO UPDATE SET "
                "major_id = EXCLUDED.major_id, title = EXCLUDED.title, "
                "title_vi = EXCLUDED.title_vi, title_en = EXCLUDED.title_en, "
                "topic_type = EXCLUDED.topic_type, status = 'ACTIVE' "
                "RETURNING id"
            ),
            {
                "semester_id": validation["semester_id"],
                "major_id": validation["major_id"],
                "code": project.code,
                "title": project.title_en or project.title_vi,
                "title_vi": project.title_vi,
                "title_en": project.title_en or None,
                "topic_type": project.topic_type,
            },
        ).scalar_one()
        if existed is None:
            created += 1
        else:
            updated += 1

        session.execute(
            text("DELETE FROM project_supervisors WHERE project_id = :project_id"),
            {"project_id": project_id},
        )
        for index, lecturer in enumerate(validation["assignments"][project.code]):
            session.execute(
                text(
                    "INSERT INTO project_supervisors "
                    "(project_id, lecturer_id, supervisor_type) "
                    "VALUES (:project_id, :lecturer_id, CAST(:supervisor_type AS supervisor_type))"
                ),
                {
                    "project_id": project_id,
                    "lecturer_id": lecturer["id"],
                    "supervisor_type": "MAIN" if index == 0 else "CO",
                },
            )
            supervisor_assignments += 1

    session.execute(
        text(
            "INSERT INTO audit_events "
            "(actor_id, action, entity_type, entity_id, reason, after_json) "
            "VALUES (NULL, 'EXCEL_PROJECTS_GVHD_IMPORTED', 'excel_import', "
            ":entity_id, :reason, CAST(:after_json AS JSONB))"
        ),
        {
            "entity_id": source_name[:96],
            "reason": "One-sheet project and supervisor import",
            "after_json": json.dumps(
                {
                    "semester": semester_code,
                    "projects": len(projects),
                    "supervisor_assignments": supervisor_assignments,
                }
            ),
        },
    )
    return {
        "semester": semester_code,
        "projects_created": created,
        "projects_updated": updated,
        "supervisor_assignments": supervisor_assignments,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--sheet", help="Force a sheet name; default finds the compatible header automatically.")
    parser.add_argument("--semester", default=DEFAULT_SEMESTER_CODE)
    parser.add_argument("--major", default=DEFAULT_MAJOR_CODE)
    parser.add_argument("--database-url", default=os.getenv("DATABASE_URL", DEFAULT_DATABASE_URL))
    parser.add_argument("--apply", action="store_true", help="Write validated projects and supervisors to PostgreSQL.")
    args = parser.parse_args()

    workbook_path = args.workbook.resolve()
    if not workbook_path.is_file():
        raise SystemExit(f"Workbook not found: {workbook_path}")

    try:
        sheet_name, projects = read_projects(workbook_path, args.sheet)
        engine = create_engine(args.database_url, pool_pre_ping=True)
        with Session(engine) as session:
            validation = validate_projects(session, projects, args.semester, args.major)
            result: dict[str, Any] = {
                "mode": "apply" if args.apply else "dry-run",
                "workbook": workbook_path.name,
                "sheet": sheet_name,
                "projects": len(projects),
                "supervisor_assignments": sum(len(project.supervisors) for project in projects),
                "existing_projects": validation["existing_projects"],
                "new_projects": validation["missing_projects"],
                "sample": [
                    {
                        "code": project.code,
                        "title_vi": project.title_vi,
                        "supervisors": list(project.supervisors),
                    }
                    for project in projects[:5]
                ],
            }
            if args.apply:
                # Validation queries open a read transaction in SQLAlchemy.
                # End it before starting the write transaction explicitly.
                session.rollback()
                with session.begin():
                    result.update(
                        apply_projects(session, projects, validation, args.semester, workbook_path.name)
                    )
            print(json.dumps(result, ensure_ascii=False, indent=2))
    except (OSError, ValueError) as exc:
        raise SystemExit(str(exc)) from exc


if __name__ == "__main__":
    main()
